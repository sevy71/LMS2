from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
from flask_migrate import Migrate
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta, timezone
import urllib.parse
from functools import wraps
from io import BytesIO
from sqlalchemy import or_

# --- Environment loading ---
# Load .env.local if it exists (for local development with Postgres connection)
# Railway production injects DATABASE_URL directly, so no .env file needed there
_env_local_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
if os.path.exists(_env_local_path):
    load_dotenv(_env_local_path)
    print(f"Loaded environment from .env.local")
else:
    # Fall back to standard .env if no .env.local
    load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config.setdefault('DISPLAY_TIMEZONE', os.environ.get('DISPLAY_TIMEZONE', 'Europe/London'))

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:  # pragma: no cover
    ZoneInfo = None

def to_local(dt: datetime) -> datetime:
    """Convert a naive/UTC datetime to configured display timezone.
    Assumes naive datetimes are UTC.
    """
    if not dt:
        return dt
    try:
        tz_name = app.config.get('DISPLAY_TIMEZONE', 'Europe/London')
        tz = ZoneInfo(tz_name) if ZoneInfo else None
        aware = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        return aware.astimezone(tz) if tz else aware
    except Exception:
        return dt

# --- Database configuration ---
def _redact_db_uri(uri: str) -> str:
    """Redact password from database URI for safe logging."""
    if not uri:
        return uri
    try:
        # Handle postgresql://user:password@host:port/db format
        if '@' in uri and ':' in uri.split('@')[0]:
            # Split into scheme://user:pass and @host/db
            prefix, rest = uri.split('@', 1)
            # Find the password portion (after last :// and before @)
            scheme_end = prefix.find('://') + 3
            user_pass = prefix[scheme_end:]
            if ':' in user_pass:
                user = user_pass.split(':')[0]
                return f"{prefix[:scheme_end]}{user}:***@{rest}"
        return uri
    except Exception:
        return "<URI redaction failed>"

# Determine which env var is providing the database URL
_db_public = os.environ.get('DATABASE_PUBLIC_URL')
_db_standard = os.environ.get('DATABASE_URL')
if _db_public:
    database_uri = _db_public
    _db_source = 'DATABASE_PUBLIC_URL'
elif _db_standard:
    database_uri = _db_standard
    _db_source = 'DATABASE_URL'
else:
    raise RuntimeError(
        "DATABASE_URL not set — refusing to start. "
        "For local development, create .env.local with DATABASE_PUBLIC_URL or DATABASE_URL. "
        "On Railway, ensure the Postgres plugin is attached."
    )

# SQLAlchemy prefers 'postgresql' over 'postgres'
app.config['SQLALCHEMY_DATABASE_URI'] = database_uri.replace('postgres://', 'postgresql://')

# Log database configuration with redacted password
print(f"[DB CONFIG] Source: {_db_source}")
print(f"[DB CONFIG] URI: {_redact_db_uri(app.config['SQLALCHEMY_DATABASE_URI'])}")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Import models and db
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from models import db, Player, Round, Fixture, Pick, PickToken, ReminderSchedule, CyclePayment


# Initialize db with app
db.init_app(app)
migrate = Migrate(app, db)

# --- Game policy configuration ---
# Postponement policy thresholds (minutes)
app.config.setdefault('POSTPONEMENT_LENIENCY_MINUTES', 60)   # early postponement window
app.config.setdefault('EARLY_PICK_WINDOW_MINUTES', 120)      # pick must predate kickoff by this to qualify

# Cycles and rounds
app.config.setdefault('MAX_ROUNDS_PER_CYCLE', 20)

# Eligibility guidance thresholds (non-blocking guidance; hard gate is >=1 eligible team)
app.config.setdefault('EARLY_ROUND_MAX', 10)
app.config.setdefault('MID_ROUND_MAX', 20)

# --- Logging helpers ---
def log_auto_pick(pick: Pick, reason: str, postponed_event_id: str = None, announcement_time: datetime = None):
    """Record that a pick was auto-assigned with policy context."""
    try:
        pick.auto_assigned = True
        pick.auto_reason = reason
        if postponed_event_id:
            pick.postponed_event_id = postponed_event_id
        if announcement_time:
            pick.announcement_time = announcement_time
        db.session.add(pick)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to log auto pick for pick_id={getattr(pick, 'id', None)}: {e}")

def set_round_special_measure(round_obj: Round, measure: str, note: str = None):
    """Apply and record a special measure on a round."""
    try:
        round_obj.special_measure = measure
        round_obj.special_note = note
        db.session.add(round_obj)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to set special measure for round_id={getattr(round_obj, 'id', None)}: {e}")

# --- Phone number sanitization ---
def sanitize_phone_number(phone_number):
    """Remove spaces, dashes, and parentheses from phone number, keeping only + and digits."""
    if not phone_number:
        return phone_number
    # Remove spaces, dashes, parentheses, and other common formatting characters
    sanitized = phone_number.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('.', '')
    return sanitized

# --- Winner detection ---
def get_current_active_round():
    """Get the current active round, prioritizing the highest cycle.
    This ensures that after a rollover, we always target the new cycle's active round.

    Returns:
        Round object or None

    Side effects:
        - Logs warning if multiple active rounds exist
        - Auto-deactivates (marks completed) any active rounds from older cycles
        - Ignores rounds with special_measure='EARLY_TERMINATED'

    Note on SEASON_BREAK:
        - Rounds with SEASON_BREAK are typically 'pending' status, not 'active'
        - This function only returns 'active' status rounds
        - To check for season break, use /api/admin/season-status or check for
          SEASON_BREAK/WAITING_FOR_FIXTURES special_measure directly
    """
    try:
        # Find all active rounds, excluding EARLY_TERMINATED ones
        # Note: SEASON_BREAK rounds typically have status='pending', not 'active',
        # so they won't be returned here. This is intentional - use season-status API
        # to detect season breaks.
        active_rounds = Round.query.filter(
            Round.status == 'active',
            or_(Round.special_measure.is_(None), Round.special_measure != 'EARLY_TERMINATED')
        ).order_by(Round.cycle_number.desc(), Round.id.desc()).all()

        if not active_rounds:
            return None

        if len(active_rounds) == 1:
            return active_rounds[0]

        # Multiple active rounds detected - select the one from the highest cycle
        current_round = active_rounds[0]  # Already ordered by cycle desc

        app.logger.warning(f"MULTIPLE ACTIVE ROUNDS DETECTED: {len(active_rounds)} active rounds found. Selecting Round {current_round.round_number} from Cycle {current_round.cycle_number}")

        # Auto-deactivate stale rounds from older cycles
        for old_round in active_rounds[1:]:
            if old_round.cycle_number < current_round.cycle_number:
                app.logger.warning(f"Auto-deactivating stale Round {old_round.round_number} from Cycle {old_round.cycle_number} (older than current Cycle {current_round.cycle_number})")
                old_round.status = 'completed'
                db.session.add(old_round)

        db.session.commit()
        return current_round

    except Exception as e:
        app.logger.error(f"Error getting current active round: {e}")
        return None

def auto_detect_and_mark_winner():
    """If exactly one active player remains, mark them as winner.
    Does nothing if zero or multiple active players remain, or if a winner is already marked.
    Returns the winner Player object if one was marked, else None.
    """
    try:
        active_players = Player.query.filter_by(status='active').all()
        if len(active_players) == 1:
            winner = active_players[0]
            if (winner.status or '').lower() != 'winner':
                winner.status = 'winner'
                db.session.add(winner)
            return winner
        return None
    except Exception as e:
        app.logger.warning(f"Winner auto-detection failed: {e}")
        return None

def handle_rollover_scenario(reference_round=None):
    """Handle rollover scenario where all remaining players lost in the same round.

    BUSINESS RULE: When zero active players remain, ALL eliminated players must be
    reactivated for a new cycle. This includes all players eliminated throughout
    the entire cycle, not just those in the final round.

    Args:
        reference_round: Optional Round object to use as reference for cycle calculation.
                        If None, uses the highest-ID active or completed round.
                        This allows rollover even if no round is marked 'completed'.

    Returns a dict with rollover info if handled, None otherwise.
    """
    try:
        # Check if we have zero active players
        active_count = Player.query.filter_by(status='active').count()
        eliminated_count = Player.query.filter_by(status='eliminated').count()

        app.logger.info(f"ROLLOVER CHECK: active={active_count}, eliminated={eliminated_count}")

        if active_count == 0 and eliminated_count > 0:
            # Determine reference round for cycle calculation
            # Priority: explicit parameter > highest active round > highest completed round
            if reference_round is None:
                reference_round = Round.query.filter_by(status='active').order_by(Round.id.desc()).first()
            if reference_round is None:
                reference_round = Round.query.filter_by(status='completed').order_by(Round.id.desc()).first()

            if not reference_round:
                app.logger.warning("ROLLOVER BLOCKED: No active or completed round found to use as reference")
                return None

            app.logger.info(f"ROLLOVER CHECK: Using Round {reference_round.round_number} (ID={reference_round.id}, status={reference_round.status}) as reference")

            # Reactivate ALL eliminated players for the new cycle
            all_eliminated_players = Player.query.filter_by(status='eliminated').all()

            if all_eliminated_players:
                app.logger.info(f"ROLLOVER TRIGGERED: Reactivating ALL {len(all_eliminated_players)} eliminated players for new cycle")

                for player in all_eliminated_players:
                    player.status = 'active'
                    db.session.add(player)

                # Calculate the next cycle number
                current_cycle = reference_round.cycle_number or 1
                next_cycle = current_cycle + 1

                # Update ALL non-completed rounds to be part of the next cycle
                # (rounds with ID > reference round that are pending or active)
                future_rounds = Round.query.filter(
                    Round.status.in_(['pending', 'active']),
                    Round.id > reference_round.id
                ).all()

                if future_rounds:
                    for round_obj in future_rounds:
                        round_obj.cycle_number = next_cycle
                        db.session.add(round_obj)
                        app.logger.info(f"ROLLOVER: Updated round {round_obj.id} (Round {round_obj.round_number}) to Cycle {next_cycle}")
                else:
                    app.logger.info("ROLLOVER: No future rounds exist yet - admin should create Round 1 of new cycle")

                db.session.commit()

                # Determine the next round number in the sequence
                next_round_num = future_rounds[0].round_number if future_rounds else (reference_round.round_number + 1)
                app.logger.info(f"ROLLOVER COMPLETE: {len(all_eliminated_players)} players reactivated for Cycle {next_cycle}")

                # AUTO-CREATE NEXT ROUND after rollover
                next_round_info = create_next_round_after_rollover(reference_round, next_cycle)
                app.logger.info(f"AUTO-CREATE NEXT ROUND: {next_round_info['message']}")

                return {
                    'handled': True,
                    'players_reactivated': len(all_eliminated_players),
                    'next_cycle': next_cycle,
                    'next_round_number': next_round_info.get('round_number') or next_round_num,
                    'reference_round_id': reference_round.id,
                    'next_round_created': next_round_info.get('created', False),
                    'next_round_id': next_round_info.get('round_id'),
                    'next_round_status': next_round_info.get('status'),
                    'next_round_fixtures': next_round_info.get('fixtures_loaded', 0),
                    'season_break': next_round_info.get('season_break', False),
                    'next_round_message': next_round_info.get('message')
                }

        return None
    except Exception as e:
        app.logger.error(f"Error handling rollover scenario: {e}")
        db.session.rollback()
        return None


# --- Season break detection and next round creation ---
# Special measure values for season handling:
#   SEASON_BREAK - No fixtures available (season ended, waiting for next season)
#   WAITING_FOR_FIXTURES - Round created but fixtures not yet available
#   EARLY_TERMINATED - Round ended early due to all players eliminated

def fetch_upcoming_fixtures(horizon_days: int = 45) -> dict:
    """Fetch upcoming Premier League fixtures to detect season availability.

    Args:
        horizon_days: Number of days to look ahead for fixtures (default 45)

    Returns:
        dict with keys:
            'available': bool - True if fixtures are available
            'fixtures_count': int - Number of upcoming fixtures found
            'next_matchday': int or None - Next available matchday
            'earliest_date': date or None - Earliest fixture date
            'error': str or None - Error message if API failed
    """
    try:
        from football_api import FootballDataAPI
        api = FootballDataAPI()

        # Fetch all fixtures for current season
        fixtures_data = api.get_premier_league_fixtures(season='2025')
        matches = fixtures_data.get('matches', [])

        if not matches:
            app.logger.warning("SEASON CHECK: No matches returned from API")
            return {
                'available': False,
                'fixtures_count': 0,
                'next_matchday': None,
                'earliest_date': None,
                'error': 'No matches returned from football API'
            }

        # Filter to upcoming scheduled matches
        now = datetime.now(timezone.utc)
        horizon_end = now + timedelta(days=horizon_days)

        upcoming = []
        for match in matches:
            if match.get('status') not in ('SCHEDULED', 'TIMED'):
                continue
            if not match.get('utcDate'):
                continue
            try:
                match_dt = datetime.fromisoformat(match['utcDate'].replace('Z', '+00:00'))
                if now <= match_dt <= horizon_end:
                    upcoming.append({
                        'matchday': match.get('matchday'),
                        'date': match_dt.date(),
                        'datetime': match_dt
                    })
            except (ValueError, TypeError):
                continue

        if not upcoming:
            app.logger.info(f"SEASON CHECK: No upcoming fixtures in next {horizon_days} days")
            return {
                'available': False,
                'fixtures_count': 0,
                'next_matchday': None,
                'earliest_date': None,
                'error': None  # Not an error - just no fixtures
            }

        # Get matchdays already used in database to avoid duplicates
        used_matchdays = {r.pl_matchday for r in Round.query.filter(Round.pl_matchday.isnot(None)).all()}
        app.logger.info(f"SEASON CHECK: Matchdays already used: {sorted(used_matchdays)}")

        # Filter out already-used matchdays
        upcoming_unused = [u for u in upcoming if u['matchday'] not in used_matchdays]

        if not upcoming_unused:
            app.logger.warning(f"SEASON CHECK: All upcoming matchdays already used, falling back to earliest")
            # Fall back to earliest fixture even if matchday is used (edge case)
            upcoming_unused = upcoming

        # Sort by date and get the earliest UNUSED matchday
        upcoming_unused.sort(key=lambda x: x['datetime'])
        next_matchday = upcoming_unused[0]['matchday']
        earliest_date = upcoming_unused[0]['date']

        app.logger.info(f"SEASON CHECK: Found {len(upcoming)} upcoming fixtures, {len(upcoming_unused)} with unused matchdays, next matchday={next_matchday}, earliest={earliest_date}")

        return {
            'available': True,
            'fixtures_count': len(upcoming),
            'next_matchday': next_matchday,
            'earliest_date': earliest_date,
            'error': None
        }

    except Exception as e:
        app.logger.error(f"SEASON CHECK failed: {e}")
        return {
            'available': False,
            'fixtures_count': 0,
            'next_matchday': None,
            'earliest_date': None,
            'error': str(e)
        }


def create_next_round_after_rollover(reference_round: Round, next_cycle: int) -> dict:
    """Create the next round automatically after a rollover.

    This function is IDEMPOTENT - if the next round already exists, it returns
    info about the existing round without creating duplicates.

    Args:
        reference_round: The round that triggered the rollover (now completed)
        next_cycle: The cycle number for the new round

    Returns:
        dict with keys:
            'created': bool - True if a new round was created
            'round_id': int - ID of the created or existing round
            'round_number': int - Round number
            'cycle_number': int - Cycle number
            'status': str - Round status
            'special_measure': str or None - Special measure if set
            'fixtures_loaded': int - Number of fixtures loaded (0 if season break)
            'season_break': bool - True if no fixtures available (season ended)
            'message': str - Human-readable result message
    """
    try:
        # Calculate the next round number (global sequence)
        max_round = db.session.query(db.func.max(Round.round_number)).scalar() or 0
        next_round_number = max_round + 1

        # Log the parameters for debugging
        app.logger.info(f"CREATE_NEXT_ROUND_AFTER_ROLLOVER: reference_round.id={reference_round.id}, "
                        f"reference_round.cycle_number={reference_round.cycle_number}, "
                        f"next_cycle={next_cycle}, next_round_number={next_round_number}")

        # IDEMPOTENCY CHECK 1: Does a round with this (round_number, cycle_number) already exist?
        existing_round = Round.query.filter_by(
            round_number=next_round_number,
            cycle_number=next_cycle
        ).first()

        if existing_round:
            app.logger.info(f"NEXT ROUND EXISTS (exact match): id={existing_round.id}, round={existing_round.round_number}, cycle={existing_round.cycle_number}")
            return {
                'created': False,
                'round_id': existing_round.id,
                'round_number': existing_round.round_number,
                'cycle_number': existing_round.cycle_number,
                'status': existing_round.status,
                'special_measure': existing_round.special_measure,
                'fixtures_loaded': len(existing_round.fixtures) if existing_round.fixtures else 0,
                'season_break': existing_round.special_measure == 'SEASON_BREAK',
                'message': f'Round {existing_round.round_number} already exists (Cycle {existing_round.cycle_number})'
            }

        # IDEMPOTENCY CHECK 2: Is there ANY active round with ID > reference_round?
        # This catches rounds created with wrong cycle_number before rollover ran
        existing_active_round = Round.query.filter(
            Round.id > reference_round.id,
            Round.status.in_(['active', 'pending']),
            or_(Round.special_measure.is_(None), Round.special_measure.notin_(['EARLY_TERMINATED']))
        ).order_by(Round.id.asc()).first()

        if existing_active_round:
            # Found a round created after reference - fix its cycle_number if wrong
            if existing_active_round.cycle_number != next_cycle:
                app.logger.warning(f"FIXING CYCLE NUMBER: Round {existing_active_round.round_number} (ID={existing_active_round.id}) "
                                   f"has cycle_number={existing_active_round.cycle_number}, should be {next_cycle}")
                existing_active_round.cycle_number = next_cycle
                db.session.add(existing_active_round)
                db.session.commit()
                app.logger.info(f"CYCLE NUMBER FIXED: Round {existing_active_round.round_number} now has cycle_number={next_cycle}")

            return {
                'created': False,
                'round_id': existing_active_round.id,
                'round_number': existing_active_round.round_number,
                'cycle_number': existing_active_round.cycle_number,
                'status': existing_active_round.status,
                'special_measure': existing_active_round.special_measure,
                'fixtures_loaded': len(existing_active_round.fixtures) if existing_active_round.fixtures else 0,
                'season_break': existing_active_round.special_measure == 'SEASON_BREAK',
                'message': f'Round {existing_active_round.round_number} already exists (cycle fixed to {next_cycle})'
            }

        # IDEMPOTENCY CHECK 3: Check for any active round in the same kickoff window
        # This prevents duplicate rounds being created for the same matchday
        existing_any_active = Round.query.filter(
            Round.status == 'active',
            or_(Round.special_measure.is_(None), Round.special_measure.notin_(['EARLY_TERMINATED', 'SEASON_BREAK']))
        ).first()
        if existing_any_active:
            app.logger.warning(f"SKIPPING ROUND CREATION: Active round already exists - "
                               f"Round {existing_any_active.round_number} (ID={existing_any_active.id}, Cycle={existing_any_active.cycle_number})")
            # Fix cycle if needed
            if existing_any_active.cycle_number != next_cycle:
                app.logger.warning(f"FIXING CYCLE NUMBER on active round: {existing_any_active.cycle_number} -> {next_cycle}")
                existing_any_active.cycle_number = next_cycle
                db.session.add(existing_any_active)
                db.session.commit()
            return {
                'created': False,
                'round_id': existing_any_active.id,
                'round_number': existing_any_active.round_number,
                'cycle_number': existing_any_active.cycle_number,
                'status': existing_any_active.status,
                'special_measure': existing_any_active.special_measure,
                'fixtures_loaded': len(existing_any_active.fixtures) if existing_any_active.fixtures else 0,
                'season_break': existing_any_active.special_measure == 'SEASON_BREAK',
                'message': f'Active round {existing_any_active.round_number} already exists (Cycle {existing_any_active.cycle_number})'
            }

        # Check fixture availability to detect season break
        fixture_check = fetch_upcoming_fixtures(horizon_days=45)

        if not fixture_check['available']:
            # SEASON BREAK: Create round in suspended state
            app.logger.info(f"SEASON BREAK DETECTED: No fixtures available")

            new_round = Round(
                round_number=next_round_number,
                cycle_number=next_cycle,
                pl_matchday=None,  # Unknown until fixtures available
                status='pending',
                special_measure='SEASON_BREAK',
                special_note='No fixtures found — season ended or break. Waiting for next season schedule. Use /api/admin/check-new-season to resume.'
            )
            db.session.add(new_round)
            db.session.commit()

            app.logger.info(f"NEXT ROUND CREATED (SEASON BREAK): id={new_round.id}, round={new_round.round_number}, cycle={new_round.cycle_number}")

            return {
                'created': True,
                'round_id': new_round.id,
                'round_number': new_round.round_number,
                'cycle_number': new_round.cycle_number,
                'status': new_round.status,
                'special_measure': new_round.special_measure,
                'fixtures_loaded': 0,
                'season_break': True,
                'message': f'Round {new_round.round_number} created but SEASON BREAK — no fixtures available'
            }

        # Fixtures available - create normal round with fixtures
        next_matchday = fixture_check['next_matchday']

        # Check if this matchday is already used in a recent round
        # If so, try the next matchday
        matchday_in_use = Round.query.filter_by(pl_matchday=next_matchday).first()
        if matchday_in_use:
            # Try to find the next available matchday
            app.logger.info(f"Matchday {next_matchday} already used, searching for next available")
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            fixtures_data = api.get_premier_league_fixtures(season='2025')

            used_matchdays = {r.pl_matchday for r in Round.query.filter(Round.pl_matchday.isnot(None)).all()}
            available_matchdays = set()
            for match in fixtures_data.get('matches', []):
                md = match.get('matchday')
                if md and md not in used_matchdays:
                    available_matchdays.add(md)

            if available_matchdays:
                next_matchday = min(available_matchdays)
                app.logger.info(f"Using next available matchday: {next_matchday}")
            else:
                # All matchdays used - season might be ending
                app.logger.warning("All matchdays used - treating as season break")
                new_round = Round(
                    round_number=next_round_number,
                    cycle_number=next_cycle,
                    pl_matchday=None,
                    status='pending',
                    special_measure='SEASON_BREAK',
                    special_note='All matchdays exhausted — waiting for next season schedule.'
                )
                db.session.add(new_round)
                db.session.commit()

                return {
                    'created': True,
                    'round_id': new_round.id,
                    'round_number': new_round.round_number,
                    'cycle_number': new_round.cycle_number,
                    'status': new_round.status,
                    'special_measure': new_round.special_measure,
                    'fixtures_loaded': 0,
                    'season_break': True,
                    'message': f'Round {new_round.round_number} created but all matchdays exhausted'
                }

        # Create the new round
        app.logger.info("=" * 60)
        app.logger.info(f">>> CREATING NEW ROUND AFTER ROLLOVER")
        app.logger.info(f"    reference_round.id={reference_round.id}")
        app.logger.info(f"    reference_round.cycle_number={reference_round.cycle_number}")
        app.logger.info(f"    computed next_cycle={next_cycle}")
        app.logger.info(f"    next_round_number={next_round_number}")
        app.logger.info(f"    next_matchday={next_matchday}")

        new_round = Round(
            round_number=next_round_number,
            cycle_number=next_cycle,
            pl_matchday=next_matchday,
            status='active',  # Ready for picks
            special_measure=None,
            special_note=f'Auto-created after rollover from Cycle {next_cycle - 1}'
        )
        db.session.add(new_round)
        db.session.flush()  # Get the ID

        app.logger.info(f"    ROUND CREATED: id={new_round.id}, cycle_number_written={new_round.cycle_number}")
        app.logger.info("=" * 60)

        # Load fixtures from API
        fixtures_loaded = 0
        earliest_kickoff = None

        try:
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            fixtures_data = api.get_premier_league_fixtures(matchday=next_matchday, season='2025')
            formatted_fixtures = api.format_fixtures_for_db(fixtures_data, next_matchday)

            for fixture_data in formatted_fixtures:
                fixture = Fixture(
                    round_id=new_round.id,
                    event_id=fixture_data['event_id'],
                    home_team=fixture_data['home_team'],
                    away_team=fixture_data['away_team'],
                    date=fixture_data['date'],
                    time=fixture_data['time'],
                    home_score=fixture_data['home_score'],
                    away_score=fixture_data['away_score'],
                    status=fixture_data['status']
                )
                db.session.add(fixture)
                fixtures_loaded += 1

                # Track earliest kickoff
                if fixture_data['date'] and fixture_data['time']:
                    try:
                        dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                        if earliest_kickoff is None or dt < earliest_kickoff:
                            earliest_kickoff = dt
                    except Exception:
                        pass

            if earliest_kickoff:
                new_round.first_kickoff_at = earliest_kickoff

        except Exception as fixture_error:
            app.logger.warning(f"Failed to load fixtures for new round: {fixture_error}")
            # Round created but no fixtures - mark as waiting
            new_round.special_measure = 'WAITING_FOR_FIXTURES'
            new_round.special_note = f'Round created but fixture loading failed: {fixture_error}'
            new_round.status = 'pending'

        db.session.commit()

        app.logger.info(f"NEXT ROUND CREATED: id={new_round.id}, round={new_round.round_number}, cycle={new_round.cycle_number}, matchday={next_matchday}, fixtures={fixtures_loaded}")

        return {
            'created': True,
            'round_id': new_round.id,
            'round_number': new_round.round_number,
            'cycle_number': new_round.cycle_number,
            'status': new_round.status,
            'special_measure': new_round.special_measure,
            'fixtures_loaded': fixtures_loaded,
            'season_break': False,
            'message': f'Round {new_round.round_number} created with {fixtures_loaded} fixtures (Matchday {next_matchday})'
        }

    except Exception as e:
        app.logger.error(f"Failed to create next round after rollover: {e}")
        db.session.rollback()
        return {
            'created': False,
            'round_id': None,
            'round_number': None,
            'cycle_number': next_cycle,
            'status': None,
            'special_measure': None,
            'fixtures_loaded': 0,
            'season_break': False,
            'message': f'Failed to create next round: {e}'
        }


# --- Optional auto-migration on startup (useful for Railway/Heroku) ---
def _auto_run_migrations_if_enabled():
    # Temporarily disabled - migration conflicts with existing database
    flag = os.environ.get('AUTO_MIGRATE', 'false').lower()
    if flag in ('1', 'true', 'yes', 'on'):
        try:
            from flask_migrate import upgrade as _upgrade
            with app.app_context():
                _upgrade()
                app.logger.info('Auto-migration completed (alembic upgrade head).')
        except Exception as e:
            app.logger.warning(f'Auto-migration failed or skipped: {e}')

# _auto_run_migrations_if_enabled()  # Disabled temporarily

# --- Fallback: Ensure required columns exist (for environments where migrations didn't run) ---
from sqlalchemy import inspect, text
from sqlalchemy.exc import OperationalError, DatabaseError

def _startup_db_ping():
    """Verify database connection at startup. Fail fast if connection fails."""
    try:
        # Simple connectivity test
        db.session.execute(text("SELECT 1"))

        # Log database version info
        dialect_name = db.engine.dialect.name
        if dialect_name == 'postgresql':
            version = db.session.execute(text("SELECT version()")).scalar()
            print(f"[DB PING] Connected to PostgreSQL: {version[:60]}...")
        elif dialect_name == 'sqlite':
            version = db.session.execute(text("SELECT sqlite_version()")).scalar()
            print(f"[DB PING] Connected to SQLite: {version}")
        else:
            print(f"[DB PING] Connected to {dialect_name}")

        db.session.rollback()  # Clean up the test transaction
        return True
    except (OperationalError, DatabaseError) as e:
        # Connection or authentication failure - fail fast
        raise RuntimeError(
            f"DATABASE CONNECTION FAILED — refusing to start.\n"
            f"URI: {_redact_db_uri(app.config['SQLALCHEMY_DATABASE_URI'])}\n"
            f"Error: {e}"
        ) from e

def _ensure_minimum_schema():
    """Ensure required columns exist (for environments where migrations didn't run).

    NOTE: This function assumes it's called within an app_context AND that the
    database connection has already been verified via _startup_db_ping().
    Schema modification failures are warnings, but connection failures should
    have been caught earlier.
    """
    try:
        engine = db.engine
        insp = inspect(engine)

        # Rounds table columns
        if insp.has_table('rounds'):
            round_cols = {col['name'] for col in insp.get_columns('rounds')}
            rounds_missing = []
            if 'first_kickoff_at' not in round_cols:
                rounds_missing.append(('first_kickoff_at', 'TIMESTAMP NULL'))
            if 'special_measure' not in round_cols:
                rounds_missing.append(('special_measure', 'VARCHAR(50) NULL'))
            if 'special_note' not in round_cols:
                rounds_missing.append(('special_note', 'TEXT NULL'))
            if 'cycle_number' not in round_cols:
                rounds_missing.append(('cycle_number', 'INTEGER NULL'))

            for name, type_sql in rounds_missing:
                try:
                    db.session.execute(text(f'ALTER TABLE rounds ADD COLUMN {name} {type_sql};'))
                    app.logger.info(f'Added missing column rounds.{name}')
                except Exception as e:
                    app.logger.warning(f'Could not add rounds.{name}: {e}')

        # Picks table columns
        if insp.has_table('picks'):
            pick_cols = {col['name'] for col in insp.get_columns('picks')}
            picks_missing = []
            if 'auto_assigned' not in pick_cols:
                picks_missing.append(('auto_assigned', 'BOOLEAN NULL'))
            if 'auto_reason' not in pick_cols:
                picks_missing.append(('auto_reason', 'VARCHAR(50) NULL'))
            if 'postponed_event_id' not in pick_cols:
                picks_missing.append(('postponed_event_id', 'VARCHAR(50) NULL'))
            if 'announcement_time' not in pick_cols:
                picks_missing.append(('announcement_time', 'TIMESTAMP NULL'))

            for name, type_sql in picks_missing:
                try:
                    db.session.execute(text(f'ALTER TABLE picks ADD COLUMN {name} {type_sql};'))
                    app.logger.info(f'Added missing column picks.{name}')
                except Exception as e:
                    app.logger.warning(f'Could not add picks.{name}: {e}')

        # Players table columns
        if insp.has_table('players'):
            player_cols = {col['name'] for col in insp.get_columns('players')}
            players_missing = []
            if 'last_entry_fee_paid_at' not in player_cols:
                players_missing.append(('last_entry_fee_paid_at', 'DATE NULL'))

            for name, type_sql in players_missing:
                try:
                    db.session.execute(text(f'ALTER TABLE players ADD COLUMN {name} {type_sql};'))
                    app.logger.info(f'Added missing column players.{name}')
                except Exception as e:
                    app.logger.warning(f'Could not add players.{name}: {e}')

        # Create reminder_schedules table if missing
        if not insp.has_table('reminder_schedules'):
            try:
                ReminderSchedule.__table__.create(bind=engine)
                app.logger.info('Created missing table reminder_schedules')
            except Exception as e:
                app.logger.warning(f'Could not create reminder_schedules: {e}')

        # Create cycle_payments table if missing (for per-cycle payment tracking)
        if not insp.has_table('cycle_payments'):
            try:
                CyclePayment.__table__.create(bind=engine)
                app.logger.info('Created missing table cycle_payments')
            except Exception as e:
                app.logger.warning(f'Could not create cycle_payments: {e}')
        else:
            # Ensure unique constraint exists (best-effort; some DBs may fail if already present)
            try:
                db.session.execute(text(
                    'CREATE UNIQUE INDEX IF NOT EXISTS uq_cycle_payment_player_cycle '
                    'ON cycle_payments (player_id, cycle_number);'
                ))
                app.logger.info('Ensured unique index on cycle_payments(player_id, cycle_number)')
            except Exception as e:
                # Constraint may already exist or DB doesn't support IF NOT EXISTS
                app.logger.debug(f'cycle_payments unique index note: {e}')

        db.session.commit()

    except (OperationalError, DatabaseError) as e:
        # Connection failures during schema ensure should not be swallowed
        db.session.rollback()
        raise RuntimeError(
            f"DATABASE ERROR during schema ensure — refusing to start.\n"
            f"Error: {e}"
        ) from e
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'Schema ensure fallback encountered an error: {e}')


# --- Startup database verification (fail-fast) ---
# Must be inside app_context and AFTER db.init_app(app)
with app.app_context():
    _startup_db_ping()       # Verify connection first - raises RuntimeError on failure
    _ensure_minimum_schema() # Then ensure schema - connection errors will also raise

# Admin authentication
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')  # Change this!
ADMIN_WHATSAPP = os.environ.get('ADMIN_WHATSAPP')  # Optional: admin WhatsApp number (e.g., +441234567890)

# --- Helpers ---
def team_abbrev(team_name: str) -> str:
    if not team_name:
        return ''

    name = team_name.strip()
    key = name.lower()
    mapping = {
        'arsenal': 'Arsenal',
        'arsenal fc': 'Arsenal',
        'aston villa': 'Villa',
        'aston villa fc': 'Villa',
        'afc bournemouth': 'Bournmouth',
        'bournemouth': 'Bournmouth',
        'bournemouth afc': 'Bournmouth',
        'brentford': 'Brentford',
        'brentford fc': 'Brentford',
        'brighton': 'Brighton',
        'brighton & hove albion': 'Brighton',
        'brighton and hove albion': 'Brighton',
        'brighton hove albion': 'Brighton',
        'burnley': 'Burnley',
        'burnley fc': 'Burnley',
        'chelsea': 'Chelsea',
        'chelsea fc': 'Chelsea',
        'crystal palace': 'Palace',
        'crystal palace fc': 'Palace',
        'palace': 'Palace',
        'everton': 'Everton',
        'everton fc': 'Everton',
        'fulham': 'Fulham',
        'fulham fc': 'Fulham',
        'leeds': 'Leeds',
        'leeds united': 'Leeds',
        'leeds united fc': 'Leeds',
        'liverpool': 'Liverpool',
        'liverpool fc': 'Liverpool',
        'manchester city': 'Man City',
        'manchester city fc': 'Man City',
        'man city': 'Man City',
        'manchester united': 'Man UTD',
        'manchester united fc': 'Man UTD',
        'man united': 'Man UTD',
        'newcastle': 'Newcastle',
        'newcastle united': 'Newcastle',
        'newcastle united fc': 'Newcastle',
        'nottingham forest': 'Forest',
        'nottm forest': 'Forest',
        'forest': 'Forest',
        'sunderland': 'Sunderland',
        'sunderland afc': 'Sunderland',
        'tottenham': 'Spurs',
        'tottenham hotspur': 'Spurs',
        'tottenham hotspur fc': 'Spurs',
        'spurs': 'Spurs',
        'west ham': 'West Ham',
        'west ham united': 'West Ham',
        'west ham united fc': 'West Ham',
        'wolverhampton wanderers': 'Wolves',
        'wolverhampton': 'Wolves',
        'wolves': 'Wolves'
    }
    return mapping.get(key, name)

def generate_picks_grid_xlsx():
    """Generate XLSX file for picks grid. Returns BytesIO object."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        rounds = Round.query.order_by(Round.round_number).all()
        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Picks Grid'

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        # Determine latest round for secondary sort
        latest_round = max(rounds, key=lambda r: r.round_number) if rounds else None

        # Sort players: Active → latest round team (A→Z, players with no pick last) → name
        def sort_key(player):
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = None
            if latest_round:
                pk = pick_map.get((player.id, latest_round.id))
                team = pk.team_picked if pk else None
            # Players with a team come first (0), then alphabetically; None teams last (1)
            team_presence = 0 if team else 1
            return (status_pri, team_presence, (team or 'zzzz'), player.name)

        for player in sorted(players, key=sort_key):
            row = [player.name, (player.status or '').upper()]
            for r in rounds:
                pick_obj = pick_map.get((player.id, r.id))
                if not pick_obj:
                    row.append('')
                else:
                    if pick_obj.is_winner is True:
                        suffix = ' (W)'
                    elif pick_obj.is_winner is False:
                        suffix = ' (L)'
                    else:
                        suffix = ' (P)'
                    row.append(f"{team_abbrev(pick_obj.team_picked)}{suffix}")
            ws.append(row)

            # Apply eliminated styling to entire row
            if (player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize columns
        for col_idx, title in enumerate(header, start=1):
            width = max(10, min(20, len(title) + 2))
            if col_idx == 1:
                width = 22
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header so sorts treat row 1 as header
        last_col_letter = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio
        
    except Exception as e:
        print(f"Error generating XLSX: {e}")
        return None

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            session.permanent = True
            next_page = request.args.get('next') or url_for('admin_dashboard')
            return redirect(next_page)
        else:
            flash('Invalid password', 'error')
    
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('index'))

@app.route('/admin/change-password', methods=['POST'])
@admin_required
def change_admin_password():
    try:
        global ADMIN_PASSWORD
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({'success': False, 'error': 'Current and new password are required'}), 400
        
        if current_password != ADMIN_PASSWORD:
            return jsonify({'success': False, 'error': 'Current password is incorrect'}), 400
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'error': 'New password must be at least 6 characters'}), 400
        
        # In a real app, you'd update the password in a database or config file
        # For now, we'll just update the environment variable (temporary)
        ADMIN_PASSWORD = new_password
        os.environ['ADMIN_PASSWORD'] = new_password
        
        return jsonify({
            'success': True,
            'message': 'Password changed successfully. Please update your ADMIN_PASSWORD environment variable in Railway.'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/picks-grid')
@admin_required
def picks_grid():
    """Display a grid of all player picks for all rounds."""
    return render_template('picks_grid.html')

@app.route('/api/picks-grid-data')
@admin_required
def get_picks_grid_data():
    """Provide data for the picks grid."""
    try:
        # Get filter parameter - default to current cycle
        cycle_filter = request.args.get('cycle', 'current')

        # Get current cycle (the highest cycle number with an active/pending round, or latest completed)
        current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.id.desc()).first()
        if not current_round:
            current_round = Round.query.order_by(Round.id.desc()).first()
        current_cycle = current_round.cycle_number or 1 if current_round else 1

        # Determine which cycles to show
        if cycle_filter == 'all':
            rounds = Round.query.order_by(Round.cycle_number, Round.round_number).all()
        elif cycle_filter == 'current':
            if current_round:
                rounds = Round.query.filter_by(cycle_number=current_cycle).order_by(Round.round_number).all()
            else:
                rounds = []
        else:
            # Explicit cycle number (e.g., "3")
            try:
                selected_cycle = int(cycle_filter)
                rounds = Round.query.filter_by(cycle_number=selected_cycle).order_by(Round.round_number).all()
            except ValueError:
                # Invalid value, fall back to current
                if current_round:
                    rounds = Round.query.filter_by(cycle_number=current_cycle).order_by(Round.round_number).all()
                else:
                    rounds = []

        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()

        # Determine the cycle number for payment lookup
        # Use selected_cycle if explicit, otherwise current_cycle
        if cycle_filter == 'all':
            payment_cycle = current_cycle  # Default to current cycle for "all" view
        elif cycle_filter == 'current':
            payment_cycle = current_cycle
        else:
            try:
                payment_cycle = int(cycle_filter)
            except ValueError:
                payment_cycle = current_cycle

        # Fetch cycle payments for the selected cycle
        cycle_payments = CyclePayment.query.filter_by(cycle_number=payment_cycle).all()
        cycle_payments_map = {cp.player_id: cp.paid_at for cp in cycle_payments}

        # Create mappings
        picks_map = {}
        results_map = {}

        for pick in picks:
            key = (pick.player_id, pick.round_id)
            picks_map[key] = pick.team_picked
            results_map[key] = {
                'is_winner': pick.is_winner,
                'is_eliminated': pick.is_eliminated
            }

        # Prepare rounds data with cycle information
        # Compute within-cycle round number for display (R1, R2, R3... per cycle)
        # round_number in DB is global sequence; we need cycle-relative position
        rounds_data = []
        cycle_round_counters = {}  # Track position within each cycle
        for r in rounds:
            cycle_num = r.cycle_number or 1
            # Increment counter for this cycle to get within-cycle position
            if cycle_num not in cycle_round_counters:
                cycle_round_counters[cycle_num] = 0
            cycle_round_counters[cycle_num] += 1
            cycle_round_number = cycle_round_counters[cycle_num]

            # Create a unique key that includes cycle info (for data lookups)
            round_key = f"C{cycle_num}-R{r.round_number}"
            # Display label uses within-cycle number (e.g., R3 instead of R11)
            display_label = f"R{cycle_round_number}" if cycle_filter != 'all' else f"C{cycle_num}-R{cycle_round_number}"
            rounds_data.append({
                'id': r.id,
                'round_number': r.round_number,
                'cycle_number': cycle_num,
                'cycle_round_number': cycle_round_number,  # Within-cycle position for display
                'round_key': round_key,
                'label': display_label
            })

        # Prepare player data
        players_data = []
        for player in players:
            player_picks = {}

            for r in rounds:
                key = (player.id, r.id)
                cycle_num = r.cycle_number or 1
                round_key = f"C{cycle_num}-R{r.round_number}"

                if key in picks_map:
                    team = picks_map[key]
                    result = results_map[key]
                    player_picks[round_key] = {
                        'team': team,
                        'is_winner': result['is_winner'],
                        'is_eliminated': result['is_eliminated']
                    }
                else:
                    player_picks[round_key] = None

            # Get cycle-specific payment date
            cycle_paid_at = cycle_payments_map.get(player.id)

            players_data.append({
                'id': player.id,
                'name': player.name,
                'status': player.status,
                'cycle_paid_at': cycle_paid_at.isoformat() if cycle_paid_at else None,
                'picks': player_picks
            })

        # Get available cycles for filtering
        all_cycles = db.session.query(Round.cycle_number).distinct().order_by(Round.cycle_number).all()
        available_cycles = [c[0] or 1 for c in all_cycles]

        return jsonify({
            'success': True,
            'rounds': rounds_data,
            'players': players_data,
            'available_cycles': available_cycles,
            'current_cycle': current_round.cycle_number or 1 if current_round else 1,
            'payment_cycle': payment_cycle,  # The cycle used for payment dates
            'cycle_filter': cycle_filter
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# DEPRECATED: This endpoint sets a global player-level payment date.
# Use /api/admin/cycles/<cycle_number>/players/<player_id>/paid-date for per-cycle payments.
@app.route('/api/admin/players/<int:player_id>/payment-date', methods=['POST'])
@admin_required
def update_player_payment_date(player_id):
    """DEPRECATED: Update the last entry fee paid date for a player (global, not per-cycle).

    This endpoint is deprecated. Use the per-cycle endpoint instead:
    POST /api/admin/cycles/<cycle_number>/players/<player_id>/paid-date
    """
    app.logger.warning(f'DEPRECATED endpoint called: /api/admin/players/{player_id}/payment-date. '
                       'Use /api/admin/cycles/<cycle>/players/<player_id>/paid-date instead.')
    try:
        player = Player.query.get(player_id)
        if not player:
            return jsonify({'success': False, 'error': 'Player not found'}), 404

        data = request.get_json()
        if data is None:
            return jsonify({'success': False, 'error': 'Invalid JSON payload'}), 400

        date_value = data.get('last_entry_fee_paid_at')

        if date_value is None or date_value == '':
            # Clear the date
            player.last_entry_fee_paid_at = None
            app.logger.info(f'Payment date cleared for player {player.name} (ID: {player_id})')
        else:
            # Parse and set the date
            try:
                parsed_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                player.last_entry_fee_paid_at = parsed_date
                app.logger.info(f'Payment date set to {parsed_date} for player {player.name} (ID: {player_id})')
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        db.session.commit()
        return jsonify({
            'success': True,
            'player_id': player_id,
            'last_entry_fee_paid_at': player.last_entry_fee_paid_at.isoformat() if player.last_entry_fee_paid_at else None
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error updating payment date for player {player_id}: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/cycles/<int:cycle_number>/players/<int:player_id>/paid-date', methods=['POST'])
@admin_required
def update_cycle_payment_date(cycle_number, player_id):
    """Update the payment date for a player in a specific cycle (upsert behavior)."""
    try:
        # Validate player exists
        player = Player.query.get(player_id)
        if not player:
            return jsonify({'success': False, 'error': 'Player not found'}), 404

        # Validate cycle_number is positive
        if cycle_number < 1:
            return jsonify({'success': False, 'error': 'Invalid cycle number'}), 400

        data = request.get_json()
        if data is None:
            return jsonify({'success': False, 'error': 'Invalid JSON payload'}), 400

        date_value = data.get('paid_at')

        # Parse the date value
        parsed_date = None
        if date_value is not None and date_value != '':
            try:
                parsed_date = datetime.strptime(date_value, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        # Upsert: find existing or create new
        cycle_payment = CyclePayment.query.filter_by(
            player_id=player_id,
            cycle_number=cycle_number
        ).first()

        if cycle_payment:
            # Update existing
            cycle_payment.paid_at = parsed_date
            app.logger.info(f'Updated cycle {cycle_number} payment for {player.name} (ID: {player_id}) to {parsed_date}')
        else:
            # Create new
            cycle_payment = CyclePayment(
                player_id=player_id,
                cycle_number=cycle_number,
                paid_at=parsed_date
            )
            db.session.add(cycle_payment)
            app.logger.info(f'Created cycle {cycle_number} payment for {player.name} (ID: {player_id}): {parsed_date}')

        db.session.commit()
        return jsonify({
            'success': True,
            'player_id': player_id,
            'cycle_number': cycle_number,
            'paid_at': cycle_payment.paid_at.isoformat() if cycle_payment.paid_at else None
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error updating cycle payment for player {player_id} cycle {cycle_number}: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin_dashboard')
@admin_required
def admin_dashboard():
    players = Player.query.all()
    current_round = get_current_active_round()
    return render_template('admin_dashboard.html', players=players, current_round=current_round)

@app.route('/api/admin/current-round-picks-status')
@admin_required
def current_round_picks_status():
    """Return pick submission status for the active round: counts, who's missing, and an optional WhatsApp link for admin when complete."""
    try:
        round_obj = get_current_active_round()
        if not round_obj:
            return jsonify({
                'success': True,
                'round': None,
                'counts': {'active_players': 0, 'picks_submitted': 0},
                'all_in': False,
                'missing': [],
                'admin_whatsapp_link': None
            })

        active_players = Player.query.filter_by(status='active').all()
        active_ids = [p.id for p in active_players]

        if not active_ids:
            return jsonify({
                'success': True,
                'round': {'id': round_obj.id, 'round_number': round_obj.round_number},
                'counts': {'active_players': 0, 'picks_submitted': 0},
                'all_in': False,
                'missing': [],
                'admin_whatsapp_link': None
            })

        picks = Pick.query.filter(Pick.round_id == round_obj.id, Pick.player_id.in_(active_ids)).all()
        picked_ids = {p.player_id for p in picks}
        missing_players = [p.name for p in active_players if p.id not in picked_ids]

        all_in = (len(picked_ids) == len(active_ids)) and len(active_ids) > 0

        # Optional WhatsApp link to notify admin when all picks are in
        whatsapp_link = None
        if all_in and ADMIN_WHATSAPP:
            base_url = os.environ.get('BASE_URL', request.url_root.rstrip('/'))
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
            if not base_url.startswith(('http://', 'https://')):
                base_url = f"https://{base_url}"

            message_lines = [
                f"✅ All picks are in!",
                f"Round {round_obj.round_number} (PL MD {round_obj.pl_matchday})",
                "",
                "You can proceed with locking the round or reviewing picks.",
                base_url
            ]
            msg = "\n".join(message_lines)
            encoded = msg.replace('\n', '%0A')
            # Sanitize and clean the admin number (remove spaces, dashes, then remove +)
            sanitized_admin = sanitize_phone_number(ADMIN_WHATSAPP)
            clean = sanitized_admin.replace('+', '')
            whatsapp_link = f"https://api.whatsapp.com/send?phone={clean}&text={encoded}"

        return jsonify({
            'success': True,
            'round': {'id': round_obj.id, 'round_number': round_obj.round_number},
            'counts': {
                'active_players': len(active_ids),
                'picks_submitted': len(picked_ids)
            },
            'all_in': all_in,
            'missing': missing_players,
            'admin_whatsapp_link': whatsapp_link
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def _earliest_kickoff_for_round(round_obj: Round):
    """Helper: determine earliest kickoff datetime for a round from fixtures."""
    try:
        earliest = None
        for fx in round_obj.fixtures or []:
            if getattr(fx, 'date', None) and getattr(fx, 'time', None):
                dt = datetime.combine(fx.date, fx.time)
                if earliest is None or dt < earliest:
                    earliest = dt
        return earliest
    except Exception:
        return None

def _eligible_teams_for_round(round_obj: Round):
    """Return the set of team names playing in this round."""
    teams = set()
    for fx in round_obj.fixtures or []:
        if fx.home_team:
            teams.add(fx.home_team)
        if fx.away_team:
            teams.add(fx.away_team)
    return teams

def _teams_used_this_cycle(player_id: int, cycle_number: int):
    """Return a set of team names the player has used in the current cycle."""
    picks = Pick.query.filter_by(player_id=player_id).join(Round).filter(Round.cycle_number == cycle_number).all()
    return {p.team_picked for p in picks}

def _opposing_team_from_past_pick(pick: Pick) -> str:
    """Find the opposing team for a given past pick, using that pick's round fixtures."""
    try:
        r = pick.round
        fixtures = r.fixtures or []
        for fx in fixtures:
            if fx.home_team == pick.team_picked:
                return fx.away_team
            if fx.away_team == pick.team_picked:
                return fx.home_team
        return None
    except Exception:
        return None

@app.route('/api/admin/rounds/<int:round_id>/apply-missed-picks', methods=['POST'])
@admin_required
def apply_missed_picks(round_id):
    """Admin-triggered: After cutoff (1h before first kickoff), auto-pick for active players without a pick.

    Logic:
    - Determine cutoff = (first_kickoff_at or derived earliest kickoff) - 1 hour.
    - For each active player with no pick in this round:
        1) Walk their past picks from most recent to oldest; when a past pick is a WIN, take the opposing (losing) team
           from that match if it's playing this round and not yet used this cycle.
        2) Otherwise, pick the first eligible team (alphabetically) that they haven't used this cycle.
    - Mark pick.auto_assigned = True, pick.auto_reason = 'missed_deadline'.
    """
    try:
        round_obj = Round.query.get_or_404(round_id)

        # Determine dry-run mode (preview only; no DB writes)
        dry_run = str(request.args.get('dry_run', 'false')).lower() in ('1', 'true', 'yes', 'y')

        # Compute cutoff time
        anchor = round_obj.first_kickoff_at or _earliest_kickoff_for_round(round_obj) or round_obj.end_date
        if not anchor:
            return jsonify({'success': False, 'error': 'Cannot determine first kickoff or deadline for this round'}), 400
        cutoff = anchor - timedelta(hours=1)
        if (datetime.utcnow() < cutoff) and (not dry_run):
            return jsonify({'success': False, 'error': 'Cutoff not reached yet. Try after the submission deadline.'}), 400

        # Build sets
        eligible_teams = _eligible_teams_for_round(round_obj)
        active_players = Player.query.filter_by(status='active').all()
        applied = []
        skipped = []

        for player in active_players:
            # Skip if player already has a pick for this round
            existing_pick = Pick.query.filter_by(player_id=player.id, round_id=round_obj.id).first()
            if existing_pick:
                skipped.append({'player': player.name, 'reason': 'already_picked'})
                continue

            used_teams = _teams_used_this_cycle(player.id, round_obj.cycle_number or 1)

            # Strategy 1: past winning picks → opposing team of that match
            candidate = None
            past_picks = Pick.query.filter_by(player_id=player.id).join(Round).order_by(Round.round_number.desc()).all()
            for past in past_picks:
                if past.is_winner is True:
                    opp = _opposing_team_from_past_pick(past)
                    if opp and (opp in eligible_teams) and (opp not in used_teams):
                        candidate = opp
                        break

            # Strategy 2: first eligible team alphabetically not yet used this cycle
            if not candidate:
                remaining = sorted([t for t in eligible_teams if t not in used_teams])
                if remaining:
                    candidate = remaining[0]

            if not candidate:
                skipped.append({'player': player.name, 'reason': 'no_eligible_team'})
                continue

            if not dry_run:
                # Create auto pick
                pick = Pick(player_id=player.id, round_id=round_obj.id, team_picked=candidate)
                db.session.add(pick)
                db.session.flush()
                # Audit
                log_auto_pick(pick, reason='missed_deadline')

            applied.append({'player': player.name, 'team': candidate})

        if not dry_run:
            db.session.commit()

        return jsonify({
            'success': True,
            'round_id': round_obj.id,
            'round_number': round_obj.round_number,
            'applied_count': len(applied),
            'applied': applied,
            'skipped': skipped,
            'dry_run': dry_run
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/send_picks')
def send_picks():
    current_round = get_current_active_round()
    if not current_round:
        return "No active round found", 404

    app.logger.info(f"Sending picks for Round {current_round.round_number}, Cycle {current_round.cycle_number}")
    active_players = Player.query.filter_by(status='active').all()
    
    for player in active_players:
        # Generate or refresh token; it will auto-expire at the round deadline if set
        pick_token = PickToken.create_for_player_round(player.id, current_round.id)
        db.session.commit() # Commit to get the token
        # Get base URL - prioritize Railway deployment URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            # Fallback to request URL but ensure it's HTTPS for production
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
        
        # Ensure base_url has protocol
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        pick_url = pick_token.get_pick_url(base_url)
        
        # Debug logging
        print(f"Generated pick URL for {player.name}: {pick_url}")
        
        # Generate general registration link
        registration_url = f"{base_url}/register"
        
        # Format message with better mobile WhatsApp compatibility
        deadline_str = current_round.end_date.strftime('%a %d %b %Y, %H:%M') if current_round.end_date else None
        message_lines = [
            f"🏆 Last Man Standing - Round {current_round.round_number}",
            "",
            f"Hi {player.name}!",
            "",
            f"Time to make your pick for Round {current_round.round_number} (PL Matchday {current_round.pl_matchday}).",
            "",
            "⚠️ Remember:",
            "• Pick a team you think will WIN",
            "• You can only use each team ONCE", 
            "• If your team loses or draws, you're out!",
            (f"• Link valid until: {deadline_str}" if deadline_str else "• Link valid until the round deadline"),
            "",
            "Good luck! 🍀",
            "",
            "Your pick link:",
            pick_url,
            "",
            "👥 Want to invite friends/family?",
            "Share this registration link:",
            registration_url
        ]
        
        message = "\n".join(message_lines)
        
        # Don't encode the URL at all - WhatsApp mobile is very sensitive to URL encoding
        # Just encode line breaks and special characters, preserve the URL completely
        encoded_message = message.replace('\n', '%0A')
        
        # Only generate WhatsApp link if player has a WhatsApp number
        if player.whatsapp_number:
            # Sanitize and clean the number (remove spaces, dashes, then remove +)
            sanitized_number = sanitize_phone_number(player.whatsapp_number)
            clean_number = sanitized_number.replace('+', '')
            # Prepare both mobile and desktop links; we will choose client-side
            player.wa_link_mobile = f"https://api.whatsapp.com/send?phone={clean_number}&text={encoded_message}"
            player.wa_link_desktop = f"https://web.whatsapp.com/send?phone={clean_number}&text={encoded_message}"
            # Backwards-compatible default (will be overridden client-side)
            player.whatsapp_link = player.wa_link_mobile
            # Debug logging
            print(f"WhatsApp links for {player.name}: mobile={player.wa_link_mobile[:80]}..., desktop={player.wa_link_desktop[:80]}...")
        else:
            player.whatsapp_link = None
            print(f"No WhatsApp number for {player.name}, link generation skipped")
        
        print(f"Pick URL in message: {pick_url}")

    return render_template('send_picks.html', players=active_players, round=current_round)

@app.route('/api/players', methods=['GET', 'POST'])
@admin_required
def handle_players():
    if request.method == 'GET':
        players = Player.query.all()
        return jsonify([{
            'id': p.id,
            'name': p.name,
            'status': p.status,
            'unreachable': p.unreachable
        } for p in players])
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            
            if not data or not data.get('name'):
                return jsonify({'success': False, 'error': 'Player name is required'}), 400
            
            # Check if player with same name already exists
            existing_player = Player.query.filter_by(name=data['name'].strip()).first()
            if existing_player:
                return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400
            
            # Create new player
            whatsapp = data.get('whatsapp_number', '').strip() or None
            player = Player(
                name=data['name'].strip(),
                whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None
            )
            
            db.session.add(player)
            db.session.commit()
            
            return jsonify({'success': True, 'id': player.id})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/bulk', methods=['POST'])
@admin_required
def bulk_import_players():
    try:
        data = request.get_json()
        
        if not data or not data.get('players'):
            return jsonify({'success': False, 'error': 'Players data is required'}), 400
        
        players_data = data['players']
        created_count = 0
        errors = []
        
        for i, player_data in enumerate(players_data):
            try:
                if not player_data.get('name'):
                    errors.append(f"Line {i+1}: Missing name")
                    continue
                
                name = player_data['name'].strip()
                whatsapp = player_data.get('whatsapp_number', '').strip()
                
                # Check if player with same name already exists
                existing_player = Player.query.filter_by(name=name).first()
                if existing_player:
                    errors.append(f"Line {i+1}: Player with name '{name}' already exists")
                    continue
                
                # WhatsApp numbers can be shared among multiple players (family members)
                # No need to check for WhatsApp duplicates anymore
                
                # Create new player
                player = Player(
                    name=name,
                    whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None
                )
                
                db.session.add(player)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Line {i+1}: {str(e)}")
        
        if created_count > 0:
            db.session.commit()
        
        if errors and created_count == 0:
            return jsonify({'success': False, 'error': 'No players could be imported', 'errors': errors}), 400
        
        response_data = {'success': True, 'count': created_count}
        if errors:
            response_data['warnings'] = errors
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/<int:player_id>', methods=['PUT', 'DELETE'])
@admin_required
def handle_player_by_id(player_id):
    player = Player.query.get_or_404(player_id)
    
    if request.method == 'PUT':
        try:
            data = request.get_json()
            
            if not data or not data.get('name'):
                return jsonify({'success': False, 'error': 'Player name is required'}), 400
            
            name = data['name'].strip()
            whatsapp = data.get('whatsapp_number', '').strip()
            
            # Check if another player with the same name exists
            existing_player = Player.query.filter(Player.name == name, Player.id != player_id).first()
            if existing_player:
                return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400
            
            # WhatsApp numbers can be shared among multiple players (family members)
            # No need to check for WhatsApp duplicates anymore
            
            # Update player
            player.name = name
            player.whatsapp_number = sanitize_phone_number(whatsapp) if whatsapp else None
            
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if player has any picks
            picks_count = Pick.query.filter_by(player_id=player_id).count()
            if picks_count > 0:
                return jsonify({'success': False, 'error': f'Cannot delete player with {picks_count} existing picks. Reset the game first to delete all picks.'}), 400

            # Delete related records in correct order to handle foreign keys
            # 1. Delete pick tokens for this player
            PickToken.query.filter_by(player_id=player_id).delete()

            # 2. Delete reminder schedules for this player
            ReminderSchedule.query.filter_by(player_id=player_id).delete()

            # 3. Now safe to delete the player
            db.session.delete(player)
            db.session.commit()

            return jsonify({'success': True})

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds', methods=['GET', 'POST'])
@admin_required
def handle_rounds():
    if request.method == 'GET':
        rounds = Round.query.all()
        return jsonify([{
            'id': r.id,
            'round_number': r.round_number,
            'status': r.status,
            'start_date': r.start_date.isoformat() if r.start_date else None,
            'end_date': r.end_date.isoformat() if r.end_date else None
        } for r in rounds])
    
    elif request.method == 'POST':
        try:
            data = request.get_json()

            # ─────────────────────────────────────────────────────────────────────
            # CYCLE NUMBER DETECTION (rollover-aware)
            # ─────────────────────────────────────────────────────────────────────
            # Check if rollover occurred since last completed round
            # by looking at the last completed round's special_measure
            last_completed = Round.query.filter_by(status='completed').order_by(Round.id.desc()).first()

            # Detect rollover state: if last completed round has EARLY_TERMINATED,
            # the next round should be in a new cycle
            rollover_occurred = False
            if last_completed and last_completed.special_measure == 'EARLY_TERMINATED':
                # Check if there's already an active round in the new cycle
                # If not, we need to increment the cycle
                existing_active = Round.query.filter(
                    Round.status.in_(['active', 'pending']),
                    Round.id > last_completed.id
                ).first()
                if not existing_active:
                    rollover_occurred = True
                    app.logger.info(f"POST /api/rounds: Detected rollover (last completed round {last_completed.round_number} is EARLY_TERMINATED)")

            # Determine current_cycle based on rollover state
            if rollover_occurred:
                # Rollover happened - new round should be in next cycle
                current_cycle = (last_completed.cycle_number or 1) + 1
                app.logger.info(f"POST /api/rounds: Using next_cycle={current_cycle} after rollover")
            else:
                # No rollover - use max cycle_number in DB
                max_cycle_row = db.session.query(db.func.max(Round.cycle_number)).scalar()
                current_cycle = max_cycle_row if max_cycle_row is not None else 1
                app.logger.info(f"POST /api/rounds: Using current_cycle={current_cycle} (no rollover detected)")

            # Auto-assign round_number if not provided
            round_number = data.get('round_number')
            if not round_number:
                # Find max round_number GLOBALLY (not per-cycle) to continue sequence
                max_round_global = Round.query.order_by(Round.round_number.desc()).first()
                round_number = (max_round_global.round_number + 1) if max_round_global else 1

            # Cycle-aware duplicate check: block only if (round_number, cycle_number) pair exists
            existing_round = Round.query.filter_by(round_number=round_number, cycle_number=current_cycle).first()
            if existing_round:
                return jsonify({'success': False, 'error': f'Round {round_number} already exists in Cycle {current_cycle}'}), 400

            # ─────────────────────────────────────────────────────────────────────
            # IDEMPOTENCY GUARD: Check if an active round already exists
            # ─────────────────────────────────────────────────────────────────────
            # Prevent creating multiple active rounds - if one already exists, block
            existing_active = Round.query.filter(
                Round.status == 'active',
                or_(Round.special_measure.is_(None), Round.special_measure.notin_(['EARLY_TERMINATED', 'SEASON_BREAK']))
            ).first()
            if existing_active:
                return jsonify({
                    'success': False,
                    'error': f'An active round already exists: Round {existing_active.round_number} (Cycle {existing_active.cycle_number}). '
                             f'Complete or deactivate it before creating a new round.'
                }), 400

            # Parse dates if provided
            start_date = None
            end_date = None

            if data.get('start_date'):
                try:
                    start_date = datetime.fromisoformat(data['start_date'].replace('T', ' '))
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid start date format'}), 400

            if data.get('end_date'):
                try:
                    end_date = datetime.fromisoformat(data['end_date'].replace('T', ' '))
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid end date format'}), 400

            # Validate date logic
            if start_date and end_date and start_date >= end_date:
                return jsonify({'success': False, 'error': 'End date must be after start date'}), 400

            # Get PL matchday
            pl_matchday = data.get('pl_matchday')
            if not pl_matchday:
                return jsonify({'success': False, 'error': 'Premier League matchday is required'}), 400

            # Create new round with explicit cycle_number
            new_round = Round(
                round_number=round_number,
                cycle_number=current_cycle,
                pl_matchday=pl_matchday,
                start_date=start_date,
                end_date=end_date,
                status=data.get('status', 'pending')
            )
            
            db.session.add(new_round)
            db.session.flush()  # Get the ID before committing
            
            # Fetch and populate fixtures
            try:
                from football_api import FootballDataAPI
                api = FootballDataAPI()
                fixtures_data = api.get_premier_league_fixtures(pl_matchday)
                formatted_fixtures = api.format_fixtures_for_db(fixtures_data, pl_matchday)
                
                if formatted_fixtures:
                    # Create fixture records from API data
                    earliest_kickoff = None
                    for fixture_data in formatted_fixtures:
                        fixture = Fixture(
                            round_id=new_round.id,
                            event_id=fixture_data['event_id'],
                            home_team=fixture_data['home_team'],
                            away_team=fixture_data['away_team'],
                            date=fixture_data['date'],
                            time=fixture_data['time'],
                            home_score=fixture_data['home_score'],
                            away_score=fixture_data['away_score'],
                            status=fixture_data['status']
                        )
                        db.session.add(fixture)
                        # Track earliest kickoff if date and time present
                        try:
                            if fixture_data['date'] and fixture_data['time']:
                                dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                                if (earliest_kickoff is None) or (dt < earliest_kickoff):
                                    earliest_kickoff = dt
                        except Exception:
                            pass
                    if earliest_kickoff:
                        new_round.first_kickoff_at = earliest_kickoff
                    
                    db.session.commit()
                    
                    return jsonify({
                        'success': True, 
                        'id': new_round.id, 
                        'round_number': new_round.round_number,
                        'pl_matchday': new_round.pl_matchday,
                        'fixtures_added': len(formatted_fixtures)
                    })
                else:
                    # No fixtures from API, create fallback fixtures
                    raise Exception("No fixtures returned from API")
                
            except Exception as fixture_error:
                print(f"API failed, creating fallback fixtures: {fixture_error}")
                # Create fallback Premier League fixtures for the round
                fallback_fixtures = [
                    ("Arsenal", "Chelsea"), ("Liverpool", "Manchester City"), 
                    ("Manchester United", "Tottenham"), ("Newcastle", "Brighton"),
                    ("Aston Villa", "West Ham"), ("Crystal Palace", "Everton"),
                    ("Fulham", "Brentford"), ("Wolves", "Nottingham Forest"),
                    ("Bournemouth", "Sheffield United"), ("Burnley", "Luton Town")
                ]
                
                for i, (home_team, away_team) in enumerate(fallback_fixtures):
                    fixture = Fixture(
                        round_id=new_round.id,
                        event_id=f"fallback_{new_round.id}_{i}",
                        home_team=home_team,
                        away_team=away_team,
                        date=None,
                        time=None,
                        home_score=None,
                        away_score=None,
                        status='scheduled'
                    )
                    db.session.add(fixture)
                
                db.session.commit()
                
                return jsonify({
                    'success': True, 
                    'id': new_round.id, 
                    'round_number': new_round.round_number,
                    'pl_matchday': new_round.pl_matchday,
                    'fixtures_added': len(fallback_fixtures),
                    'warning': f'Round created with fallback fixtures (API failed): {str(fixture_error)}'
                })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/test-matchdays')
def test_matchdays():
    """Test endpoint for debugging"""
    try:
        print("Testing matchdays endpoint...")
        matchday_data = []
        for matchday in range(1, 39):
            matchday_data.append({
                'matchday': matchday,
                'fixture_count': 10,
                'earliest_date': None,
                'latest_date': None
            })
        return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'test'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/matchdays')
@admin_required
def get_available_matchdays():
    """Get available Premier League matchdays"""
    print("=== Matchdays endpoint called ===")
    
    # Start with fallback approach to ensure it always works
    try:
        matchday_data = []
        for matchday in range(1, 39):
            matchday_data.append({
                'matchday': matchday,
                'fixture_count': 10,  # Typical PL matchday has 10 fixtures
                'earliest_date': None,
                'latest_date': None
            })
        
        print(f"Generated fallback matchdays: {len(matchday_data)} items")
        
        # Optional: Try to get real data from API if available
        try:
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            print("Attempting to get real matchday data from API...")
            
            fixtures_data = api.get_premier_league_fixtures(season='2025')
            if fixtures_data and fixtures_data.get('matches'):
                print(f"Got {len(fixtures_data['matches'])} matches from API")
                
                # Extract real matchdays
                real_matchdays = set()
                for match in fixtures_data.get('matches', []):
                    if match.get('matchday'):
                        real_matchdays.add(match['matchday'])
                
                if real_matchdays:
                    print(f"Found real matchdays: {sorted(real_matchdays)}")
                    # Replace fallback with real data
                    matchday_data = []
                    for matchday in sorted(real_matchdays):
                        fixture_count = len([m for m in fixtures_data['matches'] if m.get('matchday') == matchday])
                        matchday_data.append({
                            'matchday': matchday,
                            'fixture_count': fixture_count,
                            'earliest_date': None,
                            'latest_date': None
                        })
                    print("Using real API data")
                    return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'api'})
            
        except Exception as api_error:
            print(f"API failed (using fallback): {api_error}")
        
        # Return fallback data
        print("Using fallback matchday data")
        return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'fallback'})
        
    except Exception as e:
        print(f"Critical error in matchdays endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Matchdays endpoint failed: {str(e)}'}), 500

@app.route('/api/matchdays/<int:matchday>')
@admin_required
def get_matchday_info(matchday):
    """Get information about a specific matchday"""
    print(f"=== Getting info for matchday {matchday} ===")
    
    # Validate matchday range
    if matchday < 1 or matchday > 38:
        return jsonify({'success': False, 'error': 'Invalid matchday. Must be between 1 and 38'}), 400
    
    try:
        # Start with fallback info
        info = {
            'matchday': matchday,
            'fixture_count': 10,  # Typical PL matchday has 10 fixtures
            'earliest_date': None,
            'latest_date': None
        }
        
        # Try to get real API data to enhance the info
        try:
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            print(f"Attempting to get real data for matchday {matchday}")
            
            fixtures_data = api.get_premier_league_fixtures(matchday=matchday, season='2025')
            matches = fixtures_data.get('matches', [])
            
            if matches:
                print(f"Got {len(matches)} matches for matchday {matchday}")
                
                # Extract dates
                dates = []
                for match in matches:
                    if match.get('utcDate'):
                        try:
                            dt = datetime.fromisoformat(match['utcDate'].replace('Z', '+00:00'))
                            dates.append(dt.date())
                        except ValueError:
                            pass
                
                # Update info with real data
                info = {
                    'matchday': matchday,
                    'fixture_count': len(matches),
                    'earliest_date': min(dates).isoformat() if dates else None,
                    'latest_date': max(dates).isoformat() if dates else None
                }
                print(f"Using real API data: {info}")
                return jsonify({'success': True, 'info': info, 'source': 'api'})
            
        except Exception as api_error:
            print(f"API failed for matchday {matchday}: {api_error}")
        
        # Return fallback info
        print(f"Using fallback data for matchday {matchday}")
        return jsonify({'success': True, 'info': info, 'source': 'fallback'})
        
    except Exception as e:
        print(f"Critical error getting matchday {matchday} info: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Failed to get matchday info: {str(e)}'}), 500

@app.route('/api/rounds/<int:round_id>', methods=['GET', 'PUT', 'DELETE'])
@admin_required
def handle_round_by_id(round_id):
    """Get detailed information about a specific round, update its status, or delete it"""
    round_obj = Round.query.get_or_404(round_id)
    
    if request.method == 'GET':
        try:
            fixtures = Fixture.query.filter_by(round_id=round_id).all()
            
            return jsonify({
                'success': True,
                'round': {
                    'id': round_obj.id,
                    'round_number': round_obj.round_number,
                    'pl_matchday': round_obj.pl_matchday,
                    'status': round_obj.status,
                    'start_date': round_obj.start_date.isoformat() if round_obj.start_date else None,
                    'end_date': round_obj.end_date.isoformat() if round_obj.end_date else None,
                    'fixture_count': len(fixtures)
                }
            })
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            new_status = data.get('status')
            
            if not new_status:
                return jsonify({'success': False, 'error': 'Status is required'}), 400
                
            if new_status not in ['pending', 'active', 'completed']:
                return jsonify({'success': False, 'error': 'Invalid status. Must be pending, active, or completed'}), 400

            # If activating a round, deactivate ALL other active rounds (especially from older cycles)
            if new_status == 'active':
                other_active_rounds = Round.query.filter(
                    Round.status == 'active',
                    Round.id != round_id
                ).all()
                for old_round in other_active_rounds:
                    app.logger.warning(f"Auto-deactivating Round {old_round.round_number} (Cycle {old_round.cycle_number}) when activating Round {round_obj.round_number} (Cycle {round_obj.cycle_number})")
                    old_round.status = 'completed'
                    db.session.add(old_round)

            old_status = round_obj.status
            round_obj.status = new_status
            # If admin marks a round as completed, also attempt winner detection and handle rollover
            if new_status == 'completed':
                auto_detect_and_mark_winner()
                # Handle rollover scenario where all players were eliminated
                handle_rollover_scenario()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'round_id': round_id,
                'old_status': old_status,
                'new_status': new_status,
                'message': f'Round {round_obj.round_number} status updated from {old_status} to {new_status}'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        try:
            round_number = round_obj.round_number
            
            # Check if round has any picks
            picks_count = Pick.query.filter_by(round_id=round_id).count()
            if picks_count > 0:
                return jsonify({'success': False, 'error': f'Cannot delete round with {picks_count} existing picks'}), 400
            
            # Delete all related data in correct order (foreign key constraints)
            
            # 1. Delete pick tokens first
            tokens_deleted = PickToken.query.filter_by(round_id=round_id).delete()
            
            # 2. Delete all fixtures 
            fixtures_deleted = Fixture.query.filter_by(round_id=round_id).delete()
            
            # 3. Delete the round itself
            db.session.delete(round_obj)
            
            # Commit all deletions
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Round {round_number} deleted successfully',
                'details': {
                    'fixtures_deleted': fixtures_deleted,
                    'tokens_deleted': tokens_deleted
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/fixtures', methods=['POST'])
@admin_required
def add_fixtures_to_round(round_id):
    """Add fixtures to an existing round"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        
        # Check if round already has fixtures
        existing_fixtures = Fixture.query.filter_by(round_id=round_id).count()
        if existing_fixtures > 0:
            return jsonify({'success': False, 'error': f'Round already has {existing_fixtures} fixtures'}), 400
        
        # Try to get fixtures from API
        try:
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            fixtures_data = api.get_premier_league_fixtures(round_obj.pl_matchday)
            formatted_fixtures = api.format_fixtures_for_db(fixtures_data, round_obj.pl_matchday)
            
            if formatted_fixtures:
                # Create fixture records from API data
                earliest_kickoff = None
                for fixture_data in formatted_fixtures:
                    fixture = Fixture(
                        round_id=round_obj.id,
                        event_id=fixture_data['event_id'],
                        home_team=fixture_data['home_team'],
                        away_team=fixture_data['away_team'],
                        date=fixture_data['date'],
                        time=fixture_data['time'],
                        home_score=fixture_data['home_score'],
                        away_score=fixture_data['away_score'],
                        status=fixture_data['status']
                    )
                    db.session.add(fixture)
                    try:
                        if fixture_data['date'] and fixture_data['time']:
                            dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                            if (earliest_kickoff is None) or (dt < earliest_kickoff):
                                earliest_kickoff = dt
                    except Exception:
                        pass
                if earliest_kickoff:
                    round_obj.first_kickoff_at = earliest_kickoff
                
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'fixtures_added': len(formatted_fixtures),
                    'source': 'api'
                })
            else:
                raise Exception("No fixtures returned from API")
                
        except Exception as api_error:
            print(f"API failed, creating fallback fixtures for round {round_id}: {api_error}")
            # Create fallback Premier League fixtures
            fallback_fixtures = [
                ("Arsenal", "Chelsea"), ("Liverpool", "Manchester City"), 
                ("Manchester United", "Tottenham"), ("Newcastle", "Brighton"),
                ("Aston Villa", "West Ham"), ("Crystal Palace", "Everton"),
                ("Fulham", "Brentford"), ("Wolves", "Nottingham Forest"),
                ("Bournemouth", "Sheffield United"), ("Burnley", "Luton Town")
            ]
            
            for i, (home_team, away_team) in enumerate(fallback_fixtures):
                fixture = Fixture(
                    round_id=round_obj.id,
                    event_id=f"fallback_{round_obj.id}_{i}",
                    home_team=home_team,
                    away_team=away_team,
                    date=None,
                    time=None,
                    home_score=None,
                    away_score=None,
                    status='scheduled'
                )
                db.session.add(fixture)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'fixtures_added': len(fallback_fixtures),
                'source': 'fallback',
                'warning': f'Used fallback fixtures due to API error: {str(api_error)}'
            })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/picks')
@admin_required
def get_round_picks(round_id):
    """Get all picks and fixtures for a round"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        fixtures = Fixture.query.filter_by(round_id=round_id).all()
        picks = Pick.query.filter_by(round_id=round_id).all()
        
        # Format fixtures data
        fixtures_data = []
        for fixture in fixtures:
            fixtures_data.append({
                'id': fixture.id,
                'home_team': fixture.home_team,
                'away_team': fixture.away_team,
                'home_score': fixture.home_score,
                'away_score': fixture.away_score,
                'status': fixture.status,
                'date': fixture.date.isoformat() if fixture.date else None,
                'time': fixture.time.isoformat() if fixture.time else None
            })
        
        # Format picks data
        picks_data = []
        for pick in picks:
            picks_data.append({
                'id': pick.id,
                'player_name': pick.player.name,
                'team_picked': pick.team_picked,
                'is_winner': pick.is_winner,
                'is_eliminated': pick.is_eliminated
            })
        
        return jsonify({
            'success': True,
            'round': {
                'id': round_obj.id,
                'round_number': round_obj.round_number,
                'status': round_obj.status
            },
            'fixtures': fixtures_data,
            'picks': picks_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/auto-populate-results', methods=['POST'])
@admin_required
def auto_populate_results(round_id):
    """Auto-populate match results from the football API"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        fixtures = Fixture.query.filter_by(round_id=round_id).all()
        
        if not fixtures:
            return jsonify({'success': False, 'error': 'No fixtures found for this round'}), 400
        
        # Get updated results from API
        from football_api import FootballDataAPI
        api = FootballDataAPI()
        fixtures_data = api.get_premier_league_fixtures(round_obj.pl_matchday)
        
        if not fixtures_data or not fixtures_data.get('matches'):
            return jsonify({'success': False, 'error': 'Unable to fetch results from football API'}), 500
        
        updated_count = 0
        api_matches = fixtures_data['matches']
        
        # Update fixtures with API results
        for fixture in fixtures:
            # Find matching API fixture by team names
            for api_match in api_matches:
                if (api_match.get('homeTeam', {}).get('name') == fixture.home_team and 
                    api_match.get('awayTeam', {}).get('name') == fixture.away_team):
                    
                    # Check if match is finished and has scores
                    if api_match.get('status') == 'FINISHED':
                        score = api_match.get('score', {})
                        full_time = score.get('fullTime', {})
                        home_score = full_time.get('home')
                        away_score = full_time.get('away')
                        
                        if home_score is not None and away_score is not None:
                            fixture.home_score = home_score
                            fixture.away_score = away_score
                            fixture.status = 'completed'
                            updated_count += 1
                    break
        
        if updated_count > 0:
            db.session.commit()
            
        return jsonify({
            'success': True,
            'updated_fixtures': updated_count,
            'message': f'Updated {updated_count} fixtures with results from API'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/<int:player_id>/status', methods=['PUT'])
@admin_required
def update_player_status(player_id):
    """Manually update player status (eliminate/reactivate)"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if new_status not in ['active', 'eliminated']:
            return jsonify({'success': False, 'error': 'Invalid status. Must be "active" or "eliminated"'}), 400
        
        player = Player.query.get_or_404(player_id)
        old_status = player.status
        player.status = new_status
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'player_id': player_id,
            'player_name': player.name,
            'old_status': old_status,
            'new_status': new_status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/statistics')
@admin_required
def get_statistics():
    """Get comprehensive statistics for the competition"""
    try:
        # Competition overview stats
        total_players = Player.query.count()
        active_players = Player.query.filter_by(status='active').count()
        eliminated_players = Player.query.filter_by(status='eliminated').count()
        total_rounds = Round.query.count()
        completed_rounds = Round.query.filter_by(status='completed').count()
        active_round = Round.query.filter_by(status='active').first()
        
        # Individual player stats
        players = Player.query.all()
        player_stats = []
        
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            total_picks = len(picks)
            winning_picks = len([p for p in picks if p.is_winner])
            teams_used = list(set([p.team_picked for p in picks]))
            
            # Calculate survival streak
            survival_streak = 0
            for pick in reversed(picks):  # Start from most recent
                if pick.is_winner == True:
                    survival_streak += 1
                elif pick.is_winner == False:
                    break
            
            player_stats.append({
                'id': player.id,
                'name': player.name,
                'status': player.status,
                'total_picks': total_picks,
                'winning_picks': winning_picks,
                'success_rate': round((winning_picks / total_picks * 100) if total_picks > 0 else 0, 1),
                'teams_used': teams_used,
                'teams_used_count': len(teams_used),
                'current_streak': survival_streak
            })
        
        # Pick history for all players
        all_picks = Pick.query.join(Player).join(Round).all()
        pick_history = []
        
        for pick in all_picks:
            pick_history.append({
                'player_name': pick.player.name,
                'round_number': pick.round.round_number,
                'team_picked': pick.team_picked,
                'result': 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending'),
                'pick_date': pick.timestamp.strftime('%Y-%m-%d %H:%M') if getattr(pick, 'timestamp', None) else 'Unknown'
            })
        
        return jsonify({
            'success': True,
            'competition_stats': {
                'total_players': total_players,
                'active_players': active_players,
                'eliminated_players': eliminated_players,
                'elimination_rate': round((eliminated_players / total_players * 100) if total_players > 0 else 0, 1),
                'total_rounds': total_rounds,
                'completed_rounds': completed_rounds,
                'current_round': active_round.round_number if active_round else None
            },
            'player_stats': player_stats,
            'pick_history': pick_history
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/<export_type>')
@admin_required
def export_data(export_type):
    """Export data in CSV format"""
    try:
        import csv
        from io import StringIO
        from flask import make_response
        
        output = StringIO()
        
        if export_type == 'players':
            writer = csv.writer(output)
            writer.writerow(['ID', 'Name', 'WhatsApp Number', 'Status', 'Unreachable', 'Created Date'])
            
            players = Player.query.all()
            for player in players:
                writer.writerow([
                    player.id,
                    player.name,
                    player.whatsapp_number,
                    player.status,
                    player.unreachable,
                    player.created_at.strftime('%Y-%m-%d %H:%M:%S') if player.created_at else ''
                ])
            
            filename = 'lms_players.csv'
            
        elif export_type == 'rounds':
            writer = csv.writer(output)
            writer.writerow(['Round ID', 'Round Number', 'PL Matchday', 'Status', 'Start Date', 'End Date', 'Fixture Count'])
            
            rounds = Round.query.all()
            for round_obj in rounds:
                fixture_count = Fixture.query.filter_by(round_id=round_obj.id).count()
                writer.writerow([
                    round_obj.id,
                    round_obj.round_number,
                    round_obj.pl_matchday,
                    round_obj.status,
                    round_obj.start_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.start_date else '',
                    round_obj.end_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.end_date else '',
                    fixture_count
                ])
            
            filename = 'lms_rounds.csv'
            
        elif export_type == 'picks':
            writer = csv.writer(output)
            writer.writerow(['Pick ID', 'Player Name', 'Round Number', 'Team Picked', 'Result', 'Is Winner', 'Is Eliminated', 'Pick Date'])
            
            picks = Pick.query.join(Player).join(Round).all()
            for pick in picks:
                result = 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending')
                writer.writerow([
                    pick.id,
                    pick.player.name,
                    pick.round.round_number,
                    team_abbrev(pick.team_picked),
                    result,
                    pick.is_winner,
                    pick.is_eliminated,
                    pick.timestamp.strftime('%Y-%m-%d %H:%M:%S') if getattr(pick, 'timestamp', None) else ''
                ])
            
            filename = 'lms_picks.csv'
            
        elif export_type == 'stats':
            writer = csv.writer(output)
            writer.writerow(['Player Name', 'Status', 'Total Picks', 'Winning Picks', 'Success Rate %', 'Teams Used', 'Current Streak'])
            
            players = Player.query.all()
            for player in players:
                picks = Pick.query.filter_by(player_id=player.id).all()
                total_picks = len(picks)
                winning_picks = len([p for p in picks if p.is_winner])
                success_rate = round((winning_picks / total_picks * 100) if total_picks > 0 else 0, 1)
                teams_used = list(set([p.team_picked for p in picks]))
                
                # Calculate current streak
                survival_streak = 0
                for pick in reversed(picks):
                    if pick.is_winner == True:
                        survival_streak += 1
                    elif pick.is_winner == False:
                        break
                
                writer.writerow([
                    player.name,
                    player.status,
                    total_picks,
                    winning_picks,
                    f"{success_rate}%",
                    ', '.join(teams_used),
                    survival_streak
                ])
            
            filename = 'lms_statistics.csv'
            
        elif export_type == 'full':
            # Create a comprehensive backup with multiple sheets/sections
            writer = csv.writer(output)
            
            # Players section
            writer.writerow(['=== PLAYERS ==='])
            writer.writerow(['ID', 'Name', 'WhatsApp Number', 'Status', 'Unreachable', 'Created Date'])
            players = Player.query.all()
            for player in players:
                writer.writerow([
                    player.id, player.name, player.whatsapp_number, player.status, 
                    player.unreachable, player.created_at.strftime('%Y-%m-%d %H:%M:%S') if player.created_at else ''
                ])
            
            writer.writerow([])  # Empty row separator
            
            # Rounds section
            writer.writerow(['=== ROUNDS ==='])
            writer.writerow(['Round ID', 'Round Number', 'PL Matchday', 'Status', 'Start Date', 'End Date'])
            rounds = Round.query.all()
            for round_obj in rounds:
                writer.writerow([
                    round_obj.id, round_obj.round_number, round_obj.pl_matchday, round_obj.status,
                    round_obj.start_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.start_date else '',
                    round_obj.end_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.end_date else ''
                ])
            
            writer.writerow([])
            
            # Picks section
            writer.writerow(['=== PICKS ==='])
            writer.writerow(['Pick ID', 'Player Name', 'Round Number', 'Team Picked', 'Result', 'Pick Date'])
            picks = Pick.query.join(Player).join(Round).all()
            for pick in picks:
                result = 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending')
                writer.writerow([
                    pick.id, pick.player.name, pick.round.round_number, pick.team_picked, result,
                    pick.created_at.strftime('%Y-%m-%d %H:%M:%S') if pick.created_at else ''
                ])
            
            filename = 'lms_complete_backup.csv'
            
        else:
            return jsonify({'success': False, 'error': 'Invalid export type'}), 400
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid')
@admin_required
def export_picks_grid_csv():
    """Export a spreadsheet-style grid: Player, Status, R1..Rn with team and result."""
    try:
        import csv
        from io import StringIO
        from flask import make_response

        # Get cycle filter parameter - default to current cycle
        cycle_filter = request.args.get('cycle', 'current')

        # Determine which cycles to show (same logic as API endpoint)
        if cycle_filter == 'all':
            rounds = Round.query.order_by(Round.cycle_number, Round.round_number).all()
        else:
            # Get current cycle
            current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.id.desc()).first()
            if not current_round:
                current_round = Round.query.order_by(Round.id.desc()).first()

            if current_round:
                current_cycle = current_round.cycle_number or 1
                rounds = Round.query.filter_by(cycle_number=current_cycle).order_by(Round.round_number).all()
            else:
                rounds = []

        players = Player.query.order_by(Player.name).all()

        # Build a quick lookup for picks
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        def pick_cell(pick_obj):
            if not pick_obj:
                return ''
            if pick_obj.is_winner is True:
                suffix = ' (W)'
            elif pick_obj.is_winner is False:
                suffix = ' (L)'
            else:
                suffix = ' (P)'
            return f"{team_abbrev(pick_obj.team_picked)}{suffix}"

        output = StringIO()
        writer = csv.writer(output)

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        writer.writerow(header)

        # Rows
        for player in players:
            row = [player.name, player.status]
            for r in rounds:
                row.append(pick_cell(pick_map.get((player.id, r.id))))
            writer.writerow(row)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.csv'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks')
@admin_required
def export_round_picks_csv():
    """Export all picks for a specific round as CSV: Player, Team, Result."""
    try:
        import csv
        from io import StringIO
        from flask import make_response

        # Determine round number: query param or fallback to active or latest
        round_num_param = request.args.get('round', type=int)

        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Round', 'Player', 'Team', 'Result'])

        # Sort: Active first, then by team picked (A→Z), then by player name
        def sort_key(pick):
            player = pick.player
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = pick.team_picked or 'zzzz'
            return (status_pri, team, player.name)

        for pick in sorted(picks, key=sort_key):
            if pick.is_winner is True:
                result = 'Winner'
            elif pick.is_winner is False:
                result = 'Eliminated'
            else:
                result = 'Pending'
            writer.writerow([f"R{round_obj.round_number}", pick.player.name, team_abbrev(pick.team_picked), result])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.csv'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid-excel')
@admin_required
def export_picks_grid_excel():
    """Export a formatted Excel-compatible HTML table with eliminated rows highlighted."""
    try:
        from flask import make_response

        # Get cycle filter parameter - default to current cycle
        cycle_filter = request.args.get('cycle', 'current')

        # Determine which cycles to show (same logic as API endpoint)
        if cycle_filter == 'all':
            rounds = Round.query.order_by(Round.cycle_number, Round.round_number).all()
        else:
            # Get current cycle
            current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.id.desc()).first()
            if not current_round:
                current_round = Round.query.order_by(Round.id.desc()).first()

            if current_round:
                current_cycle = current_round.cycle_number or 1
                rounds = Round.query.filter_by(cycle_number=current_cycle).order_by(Round.round_number).all()
            else:
                rounds = []

        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        def pick_cell(pick_obj):
            if not pick_obj:
                return ''
            if pick_obj.is_winner is True:
                suffix = ' (W)'
            elif pick_obj.is_winner is False:
                suffix = ' (L)'
            else:
                suffix = ' (P)'
            return f"{pick_obj.team_picked}{suffix}"

        # Build HTML
        html = []
        html.append('<html><head><meta charset="utf-8">')
        html.append('<style>table{border-collapse:collapse;font-family:Arial,sans-serif} td,th{border:1px solid #999;padding:6px 8px} th{background:#222;color:#fff} .row-elim td{background:#f8d7da !important;color:#842029} .status-badge{padding:2px 6px;border-radius:10px;font-weight:700} .status-active{background:#198754;color:#fff} .status-eliminated{background:#dc3545;color:#fff} .status-winner{background:#0d6efd;color:#fff}</style>')
        html.append('</head><body>')
        html.append('<table>')
        # Header
        html.append('<tr><th>Player</th><th>Status</th>')
        for r in rounds:
            html.append(f'<th>R{r.round_number}</th>')
        html.append('</tr>')
        # Rows
        for player in players:
            row_class = 'row-elim' if (player.status or '').lower() == 'eliminated' else ''
            status_class = f"status-{(player.status or '').lower()}"
            status_text = (player.status or '').upper()
            html.append(f'<tr class="{row_class}"><td>{player.name}</td><td><span class="status-badge {status_class}">{status_text}</span></td>')
            for r in rounds:
                html.append(f'<td>{pick_cell(pick_map.get((player.id, r.id)))}</td>')
            html.append('</tr>')
        html.append('</table></body></html>')

        response = make_response(''.join(html))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.xls'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks-excel')
@admin_required
def export_round_picks_excel():
    """Export a formatted per-round table with eliminated rows highlighted."""
    try:
        from flask import make_response

        round_num_param = request.args.get('round', type=int)
        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        html = []
        html.append('<html><head><meta charset="utf-8">')
        html.append('<style>table{border-collapse:collapse;font-family:Arial,sans-serif} td,th{border:1px solid #999;padding:6px 8px} th{background:#222;color:#fff} .row-elim td{background:#f8d7da !important;color:#842029} .status-badge{padding:2px 6px;border-radius:10px;font-weight:700} .status-active{background:#198754;color:#fff} .status-eliminated{background:#dc3545;color:#fff} .status-winner{background:#0d6efd;color:#fff}</style>')
        html.append('</head><body>')
        html.append(f'<h3>Round {round_obj.round_number} Picks</h3>')
        html.append('<table>')
        html.append('<tr><th>Player</th><th>Status</th><th>Team</th><th>Result</th></tr>')
        for pick in picks:
            status = (pick.player.status or '').lower()
            row_class = 'row-elim' if status == 'eliminated' else ''
            if pick.is_winner is True:
                result = 'Winner'
            elif pick.is_winner is False:
                result = 'Eliminated'
            else:
                result = 'Pending'
            status_badge = f"<span class='status-badge status-{status}'>{status.upper()}</span>"
            html.append(f"<tr class='{row_class}'><td>{pick.player.name}</td><td>{status_badge}</td><td>{pick.team_picked}</td><td>{result}</td></tr>")
        html.append('</table></body></html>')

        response = make_response(''.join(html))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.xls'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid-xlsx')
@admin_required
def export_picks_grid_xlsx():
    """Export a real .xlsx workbook with eliminated rows highlighted for better compatibility (Numbers/Sheets)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from flask import make_response

        # Get cycle filter parameter - default to current cycle
        cycle_filter = request.args.get('cycle', 'current')

        # Determine which cycles to show (same logic as API endpoint)
        if cycle_filter == 'all':
            rounds = Round.query.order_by(Round.cycle_number, Round.round_number).all()
        else:
            # Get current cycle
            current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.id.desc()).first()
            if not current_round:
                current_round = Round.query.order_by(Round.id.desc()).first()

            if current_round:
                current_cycle = current_round.cycle_number or 1
                rounds = Round.query.filter_by(cycle_number=current_cycle).order_by(Round.round_number).all()
            else:
                rounds = []
        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Picks Grid'

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        # Determine latest round for secondary sort
        latest_round = max(rounds, key=lambda r: r.round_number) if rounds else None

        # Sort players: Active → latest round team (A→Z, players with no pick last) → name
        def sort_key(player):
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = None
            if latest_round:
                pk = pick_map.get((player.id, latest_round.id))
                team = pk.team_picked if pk else None
            # Players with a team come first (0), then alphabetically; None teams last (1)
            team_presence = 0 if team else 1
            return (status_pri, team_presence, (team or 'zzzz'), player.name)

        for player in sorted(players, key=sort_key):
            row = [player.name, (player.status or '').upper()]
            for r in rounds:
                pick_obj = pick_map.get((player.id, r.id))
                if not pick_obj:
                    row.append('')
                else:
                    if pick_obj.is_winner is True:
                        suffix = ' (W)'
                    elif pick_obj.is_winner is False:
                        suffix = ' (L)'
                    else:
                        suffix = ' (P)'
                    row.append(f"{team_abbrev(pick_obj.team_picked)}{suffix}")
            ws.append(row)

            # Apply eliminated styling to entire row
            if (player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize a bit
        for col_idx, title in enumerate(header, start=1):
            width = max(10, min(20, len(title) + 2))
            if col_idx == 1:
                width = 22
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header so sorts treat row 1 as header
        last_col_letter = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = make_response(bio.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.xlsx'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks-xlsx')
@admin_required
def export_round_picks_xlsx():
    """Export a per-round .xlsx with eliminated rows highlighted."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from flask import make_response

        round_num_param = request.args.get('round', type=int)
        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        wb = Workbook()
        ws = wb.active
        ws.title = f'Round {round_obj.round_number}'

        header = ['Player', 'Status', 'Team', 'Result']
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        def result_text(p):
            if p.is_winner is True:
                return 'Winner'
            if p.is_winner is False:
                return 'Eliminated'
            return 'Pending'

        # Sort: Active first, then by team picked (A→Z), then by player name
        def sort_key(pk):
            status = (pk.player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = pk.team_picked or 'zzzz'
            return (status_pri, team, pk.player.name)
        picks_sorted = sorted(picks, key=sort_key)

        for pk in picks_sorted:
            row = [pk.player.name, (pk.player.status or '').upper(), team_abbrev(pk.team_picked), result_text(pk)]
            ws.append(row)
            if (pk.player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header row
        ws.auto_filter.ref = f"A1:D{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = make_response(bio.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.xlsx'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download-export/<filename>')
def download_export_file(filename):
    """Download an exported file from the exports directory"""
    try:
        # Security: only allow downloading files from exports directory with specific pattern
        if not filename.startswith('lms_picks_grid_after_round_') or not filename.endswith('.xlsx'):
            return "File not found", 404
        
        filepath = os.path.join('exports', filename)
        if not os.path.exists(filepath):
            return "File not found", 404
        
        from flask import send_file
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return f"Error downloading file: {str(e)}", 500

@app.route('/api/rounds/<int:round_id>/process-results', methods=['POST'])
@admin_required
def process_round_results(round_id):
    """Process match results and eliminate players"""
    try:
        data = request.get_json()
        fixture_results = data.get('results', [])

        if not fixture_results:
            return jsonify({'success': False, 'error': 'No results provided'}), 400

        round_obj = Round.query.get_or_404(round_id)

        # GUARD: Block processing if round was early-terminated
        # This prevents fixture results from corrupting state after rollover
        if round_obj.special_measure == 'EARLY_TERMINATED':
            app.logger.warning(f"BLOCKED: Round {round_obj.round_number} was early-terminated. Ignoring fixture results.")
            return jsonify({
                'success': False,
                'error': f'Round {round_obj.round_number} was early-terminated (all players eliminated). '
                         f'Remaining fixtures are irrelevant and cannot be processed.',
                'special_measure': 'EARLY_TERMINATED'
            }), 400

        eliminated_players = []
        surviving_players = []

        # Update fixture results
        for result in fixture_results:
            fixture_id = result.get('fixture_id')
            home_score = result.get('home_score')  
            away_score = result.get('away_score')
            
            if fixture_id and home_score is not None and away_score is not None:
                fixture = Fixture.query.get(fixture_id)
                if fixture:
                    fixture.home_score = int(home_score)
                    fixture.away_score = int(away_score)
                    fixture.status = 'completed'
                    
                    # Determine winner/draw
                    if fixture.home_score > fixture.away_score:
                        winning_team = fixture.home_team
                    elif fixture.away_score > fixture.home_score:
                        winning_team = fixture.away_team
                    else:
                        winning_team = None  # Draw
                    
                    # Find picks for this fixture's teams
                    home_picks = Pick.query.filter_by(round_id=round_id, team_picked=fixture.home_team).all()
                    away_picks = Pick.query.filter_by(round_id=round_id, team_picked=fixture.away_team).all()
                    
                    # Process home team picks
                    for pick in home_picks:
                        if winning_team == fixture.home_team:
                            pick.is_winner = True
                            pick.is_eliminated = False
                            surviving_players.append(pick.player.name)
                        else:
                            pick.is_winner = False
                            pick.is_eliminated = True
                            pick.player.status = 'eliminated'
                            eliminated_players.append(pick.player.name)
                    
                    # Process away team picks
                    for pick in away_picks:
                        if winning_team == fixture.away_team:
                            pick.is_winner = True
                            pick.is_eliminated = False
                            surviving_players.append(pick.player.name)
                        else:
                            pick.is_winner = False
                            pick.is_eliminated = True
                            pick.player.status = 'eliminated'
                            eliminated_players.append(pick.player.name)
        
        # Check fixture completion status
        total_fixtures = Fixture.query.filter_by(round_id=round_id).count()
        completed_fixtures = Fixture.query.filter_by(round_id=round_id, status='completed').count()

        # Check current player status after processing these results
        active_players_count = Player.query.filter_by(status='active').count()
        eliminated_players_global = Player.query.filter_by(status='eliminated').count()

        # Count picks in this round and their outcomes
        round_picks = Pick.query.filter_by(round_id=round_id).all()
        picks_with_result = [p for p in round_picks if p.is_winner is not None]
        picks_winners = [p for p in round_picks if p.is_winner == True]
        picks_eliminated = [p for p in round_picks if p.is_eliminated == True]
        eliminated_picks_count = len(picks_eliminated)

        # Diagnostic logging for rollover decision
        app.logger.info("=" * 60)
        app.logger.info(f"PROCESS-RESULTS for Round {round_obj.round_number} (ID={round_id}, Cycle={round_obj.cycle_number}):")
        app.logger.info(f"  Fixtures: {completed_fixtures}/{total_fixtures} completed")
        app.logger.info(f"  Picks: {len(round_picks)} total, {len(picks_with_result)} with is_winner, {len(picks_winners)} winners, {eliminated_picks_count} eliminated")
        app.logger.info(f"  Players: {active_players_count} active, {eliminated_players_global} eliminated globally")
        app.logger.info(f"  first_kickoff_at: {round_obj.first_kickoff_at}")

        rollover_info = None
        early_termination = False

        # EARLY TERMINATION CHECK
        # Conditions (ALL must be true):
        #   1. active_players_count == 0
        #   2. eliminated_players_global > 0 (there are players to reactivate)
        #   3. Safety check: eliminated_picks_count > 0 OR (kickoff_passed AND round has picks)
        #      (prevents rollover from partial/dirty pick state before round starts)
        now_utc = datetime.now(timezone.utc)
        # Safe kickoff comparison: if first_kickoff_at is naive, assume UTC; if aware, compare directly
        if round_obj.first_kickoff_at:
            kickoff_dt = round_obj.first_kickoff_at
            if kickoff_dt.tzinfo is None:
                kickoff_dt = kickoff_dt.replace(tzinfo=timezone.utc)
            kickoff_passed = now_utc >= kickoff_dt
        else:
            kickoff_passed = False
        safe_to_terminate = eliminated_picks_count > 0 or (kickoff_passed and len(round_picks) > 0)

        if active_players_count == 0 and eliminated_players_global > 0 and safe_to_terminate:
            # Log start of early termination
            app.logger.info(f">>> EARLY TERMINATION START")
            app.logger.info(f"    active_count={active_players_count}, eliminated_count={eliminated_players_global}")
            app.logger.info(f"    reference_round_id={round_obj.id}, status={round_obj.status}, special_measure={round_obj.special_measure}")
            app.logger.info(f"    Safety: {eliminated_picks_count} eliminated picks, kickoff_passed={kickoff_passed}, round_picks={len(round_picks)}")
            app.logger.info(f"    Action: Setting status='completed' + special_measure='EARLY_TERMINATED'")
            app.logger.info(f"    Remaining {total_fixtures - completed_fixtures} fixtures are now locked out")

            # Mark round as completed AND set special_measure to prevent future processing
            round_obj.status = 'completed'
            round_obj.special_measure = 'EARLY_TERMINATED'
            round_obj.special_note = f'Early terminated: all players eliminated after {completed_fixtures}/{total_fixtures} fixtures'
            early_termination = True
            db.session.flush()

            # Handle rollover scenario - pass this round as reference
            rollover_info = handle_rollover_scenario(reference_round=round_obj)
            if rollover_info:
                reactivated_players = Player.query.filter_by(status='active').all()
                surviving_players = [p.name for p in reactivated_players]
                eliminated_players = []
                app.logger.info(f"    Rollover: SUCCESS - players_reactivated={rollover_info['players_reactivated']}, next_cycle={rollover_info['next_cycle']}")
            else:
                app.logger.warning(f"    Rollover: FAILED - handle_rollover_scenario returned None")
            app.logger.info(f">>> EARLY TERMINATION END")

        # NORMAL COMPLETION: All fixtures finished
        elif completed_fixtures == total_fixtures:
            app.logger.info(f">>> NORMAL COMPLETION: All {total_fixtures} fixtures completed")
            round_obj.status = 'completed'
            db.session.flush()

            # Check for single winner
            auto_detect_and_mark_winner()

            # Handle rollover scenario where all players were eliminated
            rollover_info = handle_rollover_scenario(reference_round=round_obj)
            if rollover_info:
                reactivated_players = Player.query.filter_by(status='active').all()
                surviving_players = [p.name for p in reactivated_players]
                eliminated_players = []
                app.logger.info(f"    Rollover: SUCCESS - {rollover_info['players_reactivated']} players reactivated for Cycle {rollover_info['next_cycle']}")
        else:
            # Round still in progress - explain why early termination didn't trigger
            app.logger.info(f">>> ROUND IN PROGRESS (no rollover)")
            if active_players_count > 0:
                app.logger.info(f"    Reason: {active_players_count} active players still remain")
            elif eliminated_players_global == 0:
                app.logger.info(f"    Reason: No eliminated players globally")
            elif not safe_to_terminate:
                app.logger.info(f"    Reason: Safety check failed (no eliminated picks and kickoff not passed)")
                app.logger.info(f"    Hint: Use POST /api/admin/run-rollover-check to trigger rollover manually")
            app.logger.info(f"    Waiting for {total_fixtures - completed_fixtures} more fixtures to complete")
        app.logger.info("=" * 60)

        db.session.commit()

        # Auto-generate XLSX file after processing results
        xlsx_file = generate_picks_grid_xlsx()
        xlsx_filename = None

        if xlsx_file:
            # Save the file to disk for direct WhatsApp sharing
            try:
                os.makedirs('exports', exist_ok=True)
                xlsx_filename = f'lms_picks_grid_after_round_{round_id}.xlsx'
                filepath = f'exports/{xlsx_filename}'
                with open(filepath, 'wb') as f:
                    f.write(xlsx_file.getvalue())
                print(f"XLSX file automatically generated after processing round {round_id} results: {filepath}")

            except Exception as e:
                print(f"Warning: Could not save XLSX file to disk: {e}")

        response_data = {
            'success': True,
            'eliminated_players': list(set(eliminated_players)),
            'surviving_players': list(set(surviving_players)),
            'total_eliminated': len(set(eliminated_players)),
            'total_surviving': len(set(surviving_players)),
            'fixtures_completed': completed_fixtures,
            'fixtures_total': total_fixtures,
            'xlsx_generated': xlsx_file is not None,
            'xlsx_filename': xlsx_filename
        }

        # Add early termination info
        if early_termination:
            response_data['early_termination'] = True
            response_data['early_termination_reason'] = f'All players eliminated after {completed_fixtures}/{total_fixtures} fixtures. Remaining fixtures skipped.'

        # Add rollover information if it was handled
        if rollover_info:
            response_data['rollover_detected'] = True
            response_data['rollover_message'] = f'Rollover scenario detected! All {rollover_info["players_reactivated"]} players have been reactivated for Cycle {rollover_info["next_cycle"]}.'
            response_data['next_round_number'] = rollover_info['next_round_number']
            response_data['next_cycle'] = rollover_info['next_cycle']

            # Include auto-created next round info
            if rollover_info.get('next_round_created'):
                response_data['next_round_created'] = True
                response_data['next_round_id'] = rollover_info.get('next_round_id')
                response_data['next_round_status'] = rollover_info.get('next_round_status')
                response_data['next_round_fixtures'] = rollover_info.get('next_round_fixtures', 0)
                response_data['next_round_message'] = rollover_info.get('next_round_message')

            # Add season break warning if applicable
            if rollover_info.get('season_break'):
                response_data['season_break'] = True
                response_data['season_break_warning'] = 'Season break detected — no fixtures available. Game suspended until next season. Use /api/admin/check-new-season to resume.'

        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/run-rollover-check', methods=['POST'])
@admin_required
def run_rollover_check():
    """Safe admin action to trigger rollover check WITHOUT requiring fixture payloads.

    This endpoint:
    1. Loads the current active round (highest ID with status='active', excluding EARLY_TERMINATED)
    2. Computes player counts (active, eliminated)
    3. If rollover conditions are met (0 active, >0 eliminated):
       - Marks the reference round: status='completed', special_measure='EARLY_TERMINATED'
       - Reactivates ALL eliminated players (global, not just those in reference round)
       - Increments cycle_number for future rounds
    4. Is IDEMPOTENT: if rollover already happened, returns success with no action

    Use this when:
    - Fixture results were updated directly in DB (not via process-results API)
    - Rollover should have triggered but didn't
    - Need to manually complete a stuck cycle

    Returns detailed diagnostics for Railway logs.
    """
    try:
        # ─────────────────────────────────────────────────────────────────────
        # STEP 1: Gather comprehensive diagnostics
        # ─────────────────────────────────────────────────────────────────────
        active_count = Player.query.filter_by(status='active').count()
        eliminated_count = Player.query.filter_by(status='eliminated').count()
        winner_count = Player.query.filter_by(status='winner').count()
        total_players = Player.query.count()

        # Use highest ID with status='active' as the reference round
        # Exclude rounds already marked EARLY_TERMINATED
        active_round = Round.query.filter(
            Round.status == 'active',
            or_(Round.special_measure.is_(None), Round.special_measure != 'EARLY_TERMINATED')
        ).order_by(Round.id.desc()).first()
        last_completed = Round.query.filter_by(status='completed').order_by(Round.id.desc()).first()

        # Log comprehensive state for Railway debugging
        app.logger.info("=" * 60)
        app.logger.info(">>> RUN-ROLLOVER-CHECK START")
        app.logger.info(f"  active_count={active_count}, eliminated_count={eliminated_count}, winner_count={winner_count}, total={total_players}")

        diagnostics = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'players': {
                'total': total_players,
                'active': active_count,
                'eliminated': eliminated_count,
                'winner': winner_count
            },
            'rounds': {
                'reference_round': None,
                'last_completed': None
            },
            'actions_taken': [],
            'rollover_triggered': False
        }

        # Determine reference round: active round (non-terminated) if exists, else last completed
        reference_round = active_round if active_round else last_completed

        if reference_round:
            # Get fixture and pick stats for reference round
            total_fixtures = Fixture.query.filter_by(round_id=reference_round.id).count()
            completed_fixtures = Fixture.query.filter_by(round_id=reference_round.id, status='completed').count()
            scheduled_fixtures = Fixture.query.filter_by(round_id=reference_round.id, status='scheduled').count()

            total_picks = Pick.query.filter_by(round_id=reference_round.id).count()
            eliminated_picks = Pick.query.filter(
                Pick.round_id == reference_round.id,
                Pick.is_eliminated == True
            ).count()
            winning_picks = Pick.query.filter(
                Pick.round_id == reference_round.id,
                Pick.is_winner == True
            ).count()

            diagnostics['rounds']['reference_round'] = {
                'id': reference_round.id,
                'round_number': reference_round.round_number,
                'cycle_number': reference_round.cycle_number,
                'status': reference_round.status,
                'special_measure': reference_round.special_measure,
                'first_kickoff_at': reference_round.first_kickoff_at.isoformat() if reference_round.first_kickoff_at else None,
                'fixtures': {
                    'total': total_fixtures,
                    'completed': completed_fixtures,
                    'scheduled': scheduled_fixtures
                },
                'picks': {
                    'total': total_picks,
                    'eliminated': eliminated_picks,
                    'winners': winning_picks
                }
            }

            app.logger.info(f"  reference_round_id={reference_round.id}, reference_round_status={reference_round.status}, reference_round_special_measure={reference_round.special_measure}")
            app.logger.info(f"  Round #{reference_round.round_number}, Cycle={reference_round.cycle_number}")
            app.logger.info(f"  first_kickoff_at: {reference_round.first_kickoff_at}")
            app.logger.info(f"  Fixtures: {completed_fixtures}/{total_fixtures} completed, {scheduled_fixtures} scheduled")
            app.logger.info(f"  Picks: {total_picks} total, {eliminated_picks} eliminated, {winning_picks} winners")

        if last_completed and last_completed != reference_round:
            diagnostics['rounds']['last_completed'] = {
                'id': last_completed.id,
                'round_number': last_completed.round_number,
                'cycle_number': last_completed.cycle_number
            }

        # ─────────────────────────────────────────────────────────────────────
        # STEP 2: IDEMPOTENCY CHECK - Has rollover already happened?
        # ─────────────────────────────────────────────────────────────────────
        # Rollover already happened if:
        # - There are active players (they were reactivated)
        # - There are no eliminated players left (all reactivated or none exist)
        # - The reference round has special_measure='EARLY_TERMINATED' AND active_count > 0
        if active_count > 0:
            reason = f'Rollover already handled or not needed: {active_count} active players exist'
            app.logger.info(f"  IDEMPOTENCY: {reason}")
            app.logger.info("=" * 60)
            return jsonify({
                'success': True,
                'rollover_triggered': False,
                'already_handled': True,
                'reason': reason,
                'diagnostics': diagnostics
            })

        if eliminated_count == 0:
            reason = f'No eliminated players to reactivate (total players: {total_players})'
            app.logger.info(f"  IDEMPOTENCY: {reason}")
            app.logger.info("=" * 60)
            return jsonify({
                'success': True,
                'rollover_triggered': False,
                'already_handled': True,
                'reason': reason,
                'diagnostics': diagnostics
            })

        # Additional idempotency: if reference round is already EARLY_TERMINATED but no active players,
        # something is wrong (rollover partially failed) - we should proceed to reactivate
        if reference_round and reference_round.special_measure == 'EARLY_TERMINATED':
            app.logger.warning(f"  Reference round already EARLY_TERMINATED but 0 active players - will reactivate")

        # ─────────────────────────────────────────────────────────────────────
        # STEP 3: Rollover conditions are met - perform rollover
        # ─────────────────────────────────────────────────────────────────────
        app.logger.info(f"  Decision: ROLLOVER NEEDED (0 active, {eliminated_count} eliminated)")

        if not reference_round:
            app.logger.error("  ERROR: No reference round found (no active or completed rounds)")
            app.logger.info("=" * 60)
            return jsonify({
                'success': False,
                'rollover_triggered': False,
                'reason': 'No active or completed round found to use as reference',
                'diagnostics': diagnostics
            }), 400

        # Mark the reference round as completed with EARLY_TERMINATED
        # This prevents future fixture processing on this round
        if reference_round.status == 'active':
            app.logger.info(f"  Action: Setting status='completed' + special_measure='EARLY_TERMINATED' on Round #{reference_round.round_number}")
            reference_round.status = 'completed'
            reference_round.special_measure = 'EARLY_TERMINATED'
            reference_round.special_note = f'Early terminated via admin run-rollover-check'
            db.session.add(reference_round)
            diagnostics['actions_taken'].append(f"Marked Round {reference_round.round_number} as completed with EARLY_TERMINATED")
        elif reference_round.special_measure != 'EARLY_TERMINATED':
            # Round is already completed but not marked as early terminated
            app.logger.info(f"  Action: Setting special_measure='EARLY_TERMINATED' on already-completed Round #{reference_round.round_number}")
            reference_round.special_measure = 'EARLY_TERMINATED'
            db.session.add(reference_round)
            diagnostics['actions_taken'].append(f"Set Round {reference_round.round_number} special_measure='EARLY_TERMINATED'")

        # Reactivate ALL eliminated players (global, not just those in reference round)
        all_eliminated = Player.query.filter_by(status='eliminated').all()
        reactivated_count = len(all_eliminated)

        app.logger.info(f"  Action: Reactivating ALL {reactivated_count} eliminated players (global)")
        for player in all_eliminated:
            player.status = 'active'
            db.session.add(player)

        # Calculate the next cycle number
        current_cycle = reference_round.cycle_number or 1
        next_cycle = current_cycle + 1

        # Update future rounds to the new cycle
        future_rounds = Round.query.filter(
            Round.status.in_(['pending', 'active']),
            Round.id > reference_round.id
        ).all()

        for future_round in future_rounds:
            app.logger.info(f"  Action: Updating Round {future_round.round_number} (ID={future_round.id}) to Cycle {next_cycle}")
            future_round.cycle_number = next_cycle
            db.session.add(future_round)

        diagnostics['actions_taken'].append(f"Reactivated {reactivated_count} players for Cycle {next_cycle}")
        if future_rounds:
            diagnostics['actions_taken'].append(f"Updated {len(future_rounds)} future rounds to Cycle {next_cycle}")

        db.session.commit()

        # Verify the new state
        new_active_count = Player.query.filter_by(status='active').count()
        diagnostics['rollover_triggered'] = True
        diagnostics['post_rollover_active_players'] = new_active_count
        diagnostics['new_cycle'] = next_cycle

        # AUTO-CREATE NEXT ROUND after rollover
        app.logger.info(f"  Action: Auto-creating next round for Cycle {next_cycle}")
        next_round_info = create_next_round_after_rollover(reference_round, next_cycle)
        app.logger.info(f"  Next Round Result: {next_round_info['message']}")

        diagnostics['next_round'] = {
            'created': next_round_info.get('created', False),
            'round_id': next_round_info.get('round_id'),
            'round_number': next_round_info.get('round_number'),
            'status': next_round_info.get('status'),
            'special_measure': next_round_info.get('special_measure'),
            'fixtures_loaded': next_round_info.get('fixtures_loaded', 0),
            'season_break': next_round_info.get('season_break', False),
            'message': next_round_info.get('message')
        }
        diagnostics['actions_taken'].append(f"Next round: {next_round_info['message']}")

        app.logger.info(f"  Result: ROLLOVER SUCCESS")
        app.logger.info(f"  players_reactivated={reactivated_count}, next_cycle={next_cycle}, new_active_count={new_active_count}")
        app.logger.info(">>> RUN-ROLLOVER-CHECK END")
        app.logger.info("=" * 60)

        response = {
            'success': True,
            'rollover_triggered': True,
            'message': f"Rollover complete! {reactivated_count} players reactivated for Cycle {next_cycle}",
            'players_reactivated': reactivated_count,
            'next_cycle': next_cycle,
            'next_round': next_round_info,
            'diagnostics': diagnostics
        }

        # Add season break warning if applicable
        if next_round_info.get('season_break'):
            response['season_break_warning'] = 'Season break detected — no fixtures available. Game suspended until next season. Use /api/admin/check-new-season to resume.'

        return jsonify(response)

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"RUN-ROLLOVER-CHECK failed with exception: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e),
            'diagnostics': diagnostics if 'diagnostics' in locals() else None
        }), 500


# Legacy endpoint - redirects to run-rollover-check for backwards compatibility
@app.route('/api/admin/force-rollover-check', methods=['POST'])
@admin_required
def force_rollover_check():
    """Legacy endpoint - redirects to run-rollover-check."""
    return run_rollover_check()


@app.route('/api/admin/check-new-season', methods=['POST'])
@admin_required
def check_new_season():
    """Check if new season fixtures are available and resume play if so.

    This endpoint should be called when the game is in SEASON_BREAK state
    to check if the new season's fixtures have become available.

    Flow:
    1. Find the most recent round with SEASON_BREAK status
    2. Check if fixtures are now available via the football API
    3. If available:
       - Update the round with fixtures
       - Set status to 'active'
       - Clear the SEASON_BREAK special_measure
    4. If not available:
       - Return info that season is still on break

    Returns JSON with:
        - success: bool
        - fixtures_available: bool
        - message: str
        - round_info: dict (if fixtures loaded)
    """
    try:
        app.logger.info("=" * 60)
        app.logger.info(">>> CHECK-NEW-SEASON START")

        # Find the most recent SEASON_BREAK round
        season_break_round = Round.query.filter(
            Round.special_measure == 'SEASON_BREAK'
        ).order_by(Round.id.desc()).first()

        # Also check for WAITING_FOR_FIXTURES rounds
        waiting_round = Round.query.filter(
            Round.special_measure == 'WAITING_FOR_FIXTURES'
        ).order_by(Round.id.desc()).first()

        # Use whichever is more recent
        target_round = None
        if season_break_round and waiting_round:
            target_round = season_break_round if season_break_round.id > waiting_round.id else waiting_round
        else:
            target_round = season_break_round or waiting_round

        if not target_round:
            app.logger.info("  No SEASON_BREAK or WAITING_FOR_FIXTURES round found")
            app.logger.info(">>> CHECK-NEW-SEASON END")
            app.logger.info("=" * 60)
            return jsonify({
                'success': True,
                'fixtures_available': False,
                'message': 'No suspended round found — game is not in season break mode',
                'round_info': None
            })

        app.logger.info(f"  Found target round: id={target_round.id}, round={target_round.round_number}, status={target_round.status}, special_measure={target_round.special_measure}")

        # Check for available fixtures
        fixture_check = fetch_upcoming_fixtures(horizon_days=45)

        if not fixture_check['available']:
            app.logger.info(f"  Fixtures still not available: {fixture_check.get('error') or 'No upcoming fixtures'}")
            app.logger.info(">>> CHECK-NEW-SEASON END")
            app.logger.info("=" * 60)
            return jsonify({
                'success': True,
                'fixtures_available': False,
                'message': 'Season still on break — no fixtures available yet. Check back later.',
                'check_result': fixture_check,
                'round_info': {
                    'id': target_round.id,
                    'round_number': target_round.round_number,
                    'status': target_round.status,
                    'special_measure': target_round.special_measure
                }
            })

        # Fixtures are available! Load them into the round
        app.logger.info(f"  Fixtures available! next_matchday={fixture_check['next_matchday']}, count={fixture_check['fixtures_count']}")

        next_matchday = fixture_check['next_matchday']

        # Check if this matchday is already used
        used_matchdays = {r.pl_matchday for r in Round.query.filter(
            Round.pl_matchday.isnot(None),
            Round.id != target_round.id  # Exclude the current round
        ).all()}

        if next_matchday in used_matchdays:
            # Find the next unused matchday
            app.logger.info(f"  Matchday {next_matchday} already used, finding next available")
            from football_api import FootballDataAPI
            api = FootballDataAPI()
            fixtures_data = api.get_premier_league_fixtures(season='2025')

            available_matchdays = set()
            for match in fixtures_data.get('matches', []):
                md = match.get('matchday')
                if md and md not in used_matchdays:
                    available_matchdays.add(md)

            if not available_matchdays:
                app.logger.warning("  All matchdays already used")
                return jsonify({
                    'success': True,
                    'fixtures_available': False,
                    'message': 'All available matchdays have been used. Season may be ending.',
                    'round_info': {
                        'id': target_round.id,
                        'round_number': target_round.round_number,
                        'status': target_round.status,
                        'special_measure': target_round.special_measure
                    }
                })

            next_matchday = min(available_matchdays)
            app.logger.info(f"  Using matchday {next_matchday}")

        # Load fixtures into the round
        from football_api import FootballDataAPI
        api = FootballDataAPI()
        fixtures_data = api.get_premier_league_fixtures(matchday=next_matchday, season='2025')
        formatted_fixtures = api.format_fixtures_for_db(fixtures_data, next_matchday)

        fixtures_loaded = 0
        earliest_kickoff = None

        # Clear any existing fixtures for this round (safety)
        Fixture.query.filter_by(round_id=target_round.id).delete()

        for fixture_data in formatted_fixtures:
            fixture = Fixture(
                round_id=target_round.id,
                event_id=fixture_data['event_id'],
                home_team=fixture_data['home_team'],
                away_team=fixture_data['away_team'],
                date=fixture_data['date'],
                time=fixture_data['time'],
                home_score=fixture_data['home_score'],
                away_score=fixture_data['away_score'],
                status=fixture_data['status']
            )
            db.session.add(fixture)
            fixtures_loaded += 1

            # Track earliest kickoff
            if fixture_data['date'] and fixture_data['time']:
                try:
                    dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                    if earliest_kickoff is None or dt < earliest_kickoff:
                        earliest_kickoff = dt
                except Exception:
                    pass

        # Update the round
        target_round.pl_matchday = next_matchday
        target_round.status = 'active'
        target_round.special_measure = None
        target_round.special_note = f'Resumed from season break. Matchday {next_matchday} loaded.'
        if earliest_kickoff:
            target_round.first_kickoff_at = earliest_kickoff

        db.session.commit()

        app.logger.info(f"  Round {target_round.round_number} resumed with {fixtures_loaded} fixtures from Matchday {next_matchday}")
        app.logger.info(">>> CHECK-NEW-SEASON END")
        app.logger.info("=" * 60)

        return jsonify({
            'success': True,
            'fixtures_available': True,
            'message': f'Season resumed! Round {target_round.round_number} activated with {fixtures_loaded} fixtures from Matchday {next_matchday}',
            'round_info': {
                'id': target_round.id,
                'round_number': target_round.round_number,
                'cycle_number': target_round.cycle_number,
                'status': target_round.status,
                'pl_matchday': target_round.pl_matchday,
                'fixtures_loaded': fixtures_loaded,
                'first_kickoff_at': target_round.first_kickoff_at.isoformat() if target_round.first_kickoff_at else None
            }
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"CHECK-NEW-SEASON failed: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/season-status', methods=['GET'])
@admin_required
def get_season_status():
    """Get the current season status including any breaks or suspensions.

    Returns:
        - season_active: bool - True if game is running normally
        - season_break: bool - True if in season break
        - waiting_for_fixtures: bool - True if waiting for fixtures
        - current_round: dict - Info about the current/latest round
        - fixture_availability: dict - Result from fetch_upcoming_fixtures()
        - message: str - Human-readable status
    """
    try:
        # Check for SEASON_BREAK or WAITING_FOR_FIXTURES rounds
        season_break_round = Round.query.filter(
            Round.special_measure == 'SEASON_BREAK'
        ).order_by(Round.id.desc()).first()

        waiting_round = Round.query.filter(
            Round.special_measure == 'WAITING_FOR_FIXTURES'
        ).order_by(Round.id.desc()).first()

        # Get current active round
        active_round = get_current_active_round()

        # Check fixture availability
        fixture_check = fetch_upcoming_fixtures(horizon_days=45)

        # Determine status
        season_break = season_break_round is not None
        waiting_for_fixtures = waiting_round is not None
        season_active = active_round is not None and not season_break and not waiting_for_fixtures

        # Build current round info
        current_round = None
        if season_break_round:
            current_round = {
                'id': season_break_round.id,
                'round_number': season_break_round.round_number,
                'status': season_break_round.status,
                'special_measure': season_break_round.special_measure,
                'special_note': season_break_round.special_note
            }
        elif waiting_round:
            current_round = {
                'id': waiting_round.id,
                'round_number': waiting_round.round_number,
                'status': waiting_round.status,
                'special_measure': waiting_round.special_measure,
                'special_note': waiting_round.special_note
            }
        elif active_round:
            current_round = {
                'id': active_round.id,
                'round_number': active_round.round_number,
                'status': active_round.status,
                'special_measure': active_round.special_measure,
                'pl_matchday': active_round.pl_matchday,
                'first_kickoff_at': active_round.first_kickoff_at.isoformat() if active_round.first_kickoff_at else None
            }

        # Build message
        if season_break:
            message = 'Season break — game suspended until fixtures return. Use "Check for new season" to resume.'
        elif waiting_for_fixtures:
            message = 'Waiting for fixtures — round created but fixtures not yet loaded.'
        elif season_active:
            message = f'Season active — Round {active_round.round_number} is in progress.'
        else:
            message = 'No active round. Create a new round to continue.'

        return jsonify({
            'success': True,
            'season_active': season_active,
            'season_break': season_break,
            'waiting_for_fixtures': waiting_for_fixtures,
            'current_round': current_round,
            'fixture_availability': fixture_check,
            'message': message
        })

    except Exception as e:
        app.logger.error(f"GET-SEASON-STATUS failed: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/import-historical-picks', methods=['POST'])
@admin_required
def import_historical_picks():
    """Import historical picks for rounds 1 and 2"""
    try:
        # Historical data with name mapping for exact matches
        historical_data = {
            1: {  # Round 1 picks - using exact database names
                "A. Frost": ("Liverpool", True),
                "Andy Urmson": ("Liverpool", True), 
                "Chris Hollows": ("Spurs", True),
                "Dan Groves": ("Liverpool", True),
                "Greg Leigh": ("Liverpool", True),
                "Jimmy Winning": ("Liverpool", True),
                "Mrs Shooter": ("Leeds", True),
                "P. Warby": ("Forest", True),
                "Rich Amis": ("Liverpool", True),
                "Stu Hall": ("Spurs", True),
                "Terry Leigh": ("Liverpool", True),
                "Vicky Hughes": ("Spurs", True)
            },
            2: {  # Round 2 picks - using exact database names
                "A. Frost": ("Arsenal", True),
                "Andy Urmson": ("Arsenal", True),
                "Chris Hollows": ("Arsenal", True), 
                "Dan Groves": ("Arsenal", True),
                "Greg Leigh": ("Arsenal", True),
                "Jimmy Winning": ("Chelsea", True),
                "Mrs Shooter": ("Liverpool", True),
                "P. Warby": ("Arsenal", True),
                "Rich Amis": ("Arsenal", True),
                "Stu Hall": ("Chelsea", True),
                "Terry Leigh": ("Arsenal", True),
                "Vicky Hughes": ("Arsenal", True)
            }
        }
        
        imported_count = 0
        not_found_players = []
        already_exists = []
        
        for round_num, picks_data in historical_data.items():
            round_obj = Round.query.filter_by(round_number=round_num).first()
            if not round_obj:
                continue
                
            for player_name, (team, is_winner) in picks_data.items():
                player = Player.query.filter_by(name=player_name).first()
                if not player:
                    not_found_players.append(player_name)
                    continue
                
                # Check if pick already exists using raw SQL
                result = db.session.execute(db.text(
                    "SELECT id FROM picks WHERE player_id = :player_id AND round_id = :round_id"
                ), {"player_id": player.id, "round_id": round_obj.id})
                
                if result.fetchone():
                    already_exists.append(f"{player_name} R{round_num}")
                    continue
                
                # Create the pick using raw SQL to avoid schema issues
                db.session.execute(db.text(
                    "INSERT INTO picks (player_id, round_id, team_picked, is_winner, is_eliminated) "
                    "VALUES (:player_id, :round_id, :team_picked, :is_winner, :is_eliminated)"
                ), {
                    "player_id": player.id,
                    "round_id": round_obj.id, 
                    "team_picked": team,
                    "is_winner": is_winner,
                    "is_eliminated": False if is_winner else (True if is_winner is False else False)
                })
                imported_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {imported_count} historical picks',
            'imported_count': imported_count,
            'not_found_players': list(set(not_found_players)),
            'already_exists': already_exists
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/debug-used-teams/<int:player_id>')
@admin_required  
def debug_used_teams(player_id):
    """Debug endpoint to check which teams a player has used"""
    try:
        player = Player.query.get_or_404(player_id)
        
        # Use raw SQL to get picks
        result = db.session.execute(db.text(
            "SELECT r.round_number, picks.team_picked, picks.is_winner "
            "FROM picks JOIN rounds r ON picks.round_id = r.id "
            "WHERE picks.player_id = :player_id ORDER BY r.round_number"
        ), {"player_id": player_id})
        
        picks_data = result.fetchall()
        used_teams = [row[1] for row in picks_data]
        
        # Get current round fixtures to compare team names
        current_round = Round.query.filter_by(status='active').first()
        fixture_teams = []
        if current_round:
            fixture_result = db.session.execute(db.text(
                "SELECT home_team, away_team FROM fixtures WHERE round_id = :round_id"
            ), {"round_id": current_round.id})
            
            for home, away in fixture_result.fetchall():
                fixture_teams.extend([home, away])
        
        return jsonify({
            'success': True,
            'player_name': player.name,
            'player_status': player.status,
            'picks_history': [
                {
                    'round': row[0],
                    'team': row[1], 
                    'result': 'WIN' if row[2] else 'LOSE' if row[2] is False else 'PENDING'
                } for row in picks_data
            ],
            'used_teams': used_teams,
            'fixture_teams': list(set(fixture_teams)),
            'team_matches': {team: team in fixture_teams for team in used_teams},
            'total_picks': len(picks_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/emergency-delete-round4', methods=['POST'])
@admin_required
def emergency_delete_round4():
    """Emergency endpoint to delete Round 4 with all related data"""
    try:
        # Use raw SQL to avoid model issues
        result = db.session.execute(db.text("SELECT id FROM rounds WHERE round_number = 4"))
        round4_row = result.fetchone()
        
        if not round4_row:
            return jsonify({'success': False, 'error': 'Round 4 not found'}), 404
        
        round4_id = round4_row[0]
        
        # Delete in correct order using raw SQL
        db.session.execute(db.text("DELETE FROM pick_tokens WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM picks WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM fixtures WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM rounds WHERE id = :round_id"), {"round_id": round4_id})
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Round 4 emergency deleted successfully using raw SQL'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reset-game', methods=['POST'])
@admin_required
def reset_game():
    """Reset the game by deleting all game data except players"""
    try:
        # Count items before deletion for reporting
        rounds_count = Round.query.count()
        fixtures_count = Fixture.query.count()
        picks_count = Pick.query.count()
        pick_tokens_count = PickToken.query.count()
        reminder_schedules_count = ReminderSchedule.query.count()
        players_count = Player.query.count()

        # Delete in correct order to handle foreign key constraints
        # Use synchronize_session=False for PostgreSQL compatibility

        # 1. Delete pick tokens (references players and rounds)
        PickToken.query.delete(synchronize_session=False)

        # 2. Delete all picks (references players and rounds)
        Pick.query.delete(synchronize_session=False)

        # 3. Delete reminder schedules (references players and rounds)
        ReminderSchedule.query.delete(synchronize_session=False)

        # 4. Delete all fixtures (references rounds)
        Fixture.query.delete(synchronize_session=False)

        # 5. Delete all rounds (now safe to delete)
        Round.query.delete(synchronize_session=False)

        # 6. Reset all players to active status (but keep the player records)
        Player.query.update({'status': 'active', 'unreachable': False}, synchronize_session=False)

        # Commit all changes
        db.session.commit()

        return jsonify({
            'success': True,
            'rounds_deleted': rounds_count,
            'fixtures_deleted': fixtures_count,
            'picks_deleted': picks_count,
            'pick_tokens_deleted': pick_tokens_count,
            'reminder_schedules_deleted': reminder_schedules_count,
            'players_reset': players_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/pick/<token>', methods=['GET', 'POST'])
def make_pick(token):
    # Find the pick token
    pick_token = PickToken.query.filter_by(token=token).first()
    
    if not pick_token:
        return render_template('pick_error.html', error="Invalid pick link", player_nav_only=True), 404
    
    if not pick_token.is_valid():
        error = "This pick link has expired" if pick_token.expires_at and datetime.utcnow() > pick_token.expires_at else "This pick link has already been used"
        return render_template('pick_error.html', error=error, player_nav_only=True), 400
    
    player = pick_token.player
    round_obj = pick_token.round
    
    # Check if player already has a pick for this round
    existing_pick = Pick.query.filter_by(player_id=player.id, round_id=round_obj.id).first()
    can_edit = pick_token.edit_count < 2
    edits_remaining = 2 - pick_token.edit_count
    
    # If pick exists but token has no more edits, show read-only success page
    if existing_pick and not can_edit:
        return render_template('pick_success.html', 
                             player=player, 
                             round=round_obj, 
                             team_picked=existing_pick.team_picked,
                             already_picked=True,
                             can_edit=False,
                             edits_remaining=0,
                             token=token,
                             player_nav_only=True)
    
    # Get fixtures for this round
    fixtures = Fixture.query.filter_by(round_id=round_obj.id).all()
    print(f"Found {len(fixtures)} fixtures for round {round_obj.id} (round number {round_obj.round_number})")

    # If no fixtures exist, this indicates a problem with round creation
    if not fixtures:
        print(f"ERROR: No fixtures found for round {round_obj.id}. This round may have been created without fixtures.")

    # Get player's previous picks for THIS CYCLE ONLY to prevent reusing teams
    # This ensures teams become available again after a rollover (new cycle)
    current_cycle = round_obj.cycle_number or 1
    previous_picks = Pick.query.filter_by(player_id=player.id).join(Round).filter(
        Round.cycle_number == current_cycle
    ).all()
    used_teams = [pick.team_picked for pick in previous_picks]

    print(f"Player {player.name} in Cycle {current_cycle}: {len(used_teams)} teams used this cycle")
    
    # Create a normalized team matching function to handle name variations
    def normalize_team_name(team_name):
        """Normalize team names for comparison"""
        if not team_name:
            return ""
        # Remove common suffixes and normalize
        normalized = team_name.lower()
        normalized = normalized.replace(' fc', '').replace(' afc', '').replace(' united fc', '')
        normalized = normalized.replace('tottenham hotspur', 'spurs').replace('nottingham forest', 'forest')
        normalized = normalized.replace('wolverhampton wanderers', 'wolves')
        normalized = normalized.replace('brighton & hove albion', 'brighton')
        normalized = normalized.replace('afc bournemouth', 'bournemouth')
        normalized = normalized.replace('west ham united', 'west ham')
        return normalized.strip()
    
    # Create a set of normalized used team names for faster lookup
    normalized_used_teams = {normalize_team_name(team) for team in used_teams}
    
    # Function to check if a team is already used
    def is_team_used(fixture_team_name):
        return normalize_team_name(fixture_team_name) in normalized_used_teams
    
    # Debug logging for team availability
    all_teams = []
    for fixture in fixtures:
        all_teams.extend([fixture.home_team, fixture.away_team])
    available_teams = [team for team in all_teams if team not in used_teams]
    
    print(f"Player {player.name}: {len(used_teams)} used teams, {len(available_teams)} available teams")
    print(f"Used teams: {used_teams}")
    print(f"Available teams: {set(available_teams)}")
    
    if request.method == 'POST':
        team_picked = request.form.get('team_picked')
        
        if not team_picked:
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="Please select a team",
                                 player_nav_only=True)
        
        if is_team_used(team_picked):
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="You have already picked this team in a previous round",
                                 player_nav_only=True)
        
        # Validate team exists in fixtures
        valid_teams = []
        for fixture in fixtures:
            valid_teams.extend([fixture.home_team, fixture.away_team])
        
        if team_picked not in valid_teams:
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="Invalid team selection",
                                 player_nav_only=True)
        
        # Create or update the pick
        if existing_pick:
            # Update existing pick
            existing_pick.team_picked = team_picked
            existing_pick.last_edited_at = datetime.utcnow()
            is_new_pick = False
        else:
            # Create new pick
            pick = Pick(
                player_id=player.id,
                round_id=round_obj.id,
                team_picked=team_picked
            )
            db.session.add(pick)
            is_new_pick = True
        
        pick_token.mark_used()
        db.session.commit()
        
        return render_template('pick_success.html', 
                             player=player, 
                             round=round_obj, 
                             team_picked=team_picked,
                             already_picked=not is_new_pick,
                             can_edit=pick_token.edit_count < 2,
                             edits_remaining=2 - pick_token.edit_count,
                             token=token,
                             player_nav_only=True)
    
    # GET request - show the pick form
    return render_template('pick_form.html', 
                         player=player, 
                         round=round_obj, 
                         fixtures=fixtures, 
                         used_teams=used_teams,
                         existing_pick=existing_pick,
                         can_edit=can_edit,
                         edits_remaining=edits_remaining,
                         token=token,
                         is_team_used=is_team_used,
                         player_nav_only=True)

@app.route('/register')
def player_registration():
    """Show player registration form for existing players to invite family members"""
    return render_template('player_registration.html')

@app.route('/register/<whatsapp_number>')
def register_with_whatsapp(whatsapp_number):
    """Show registration form pre-filled with WhatsApp number"""
    # Decode the whatsapp number (in case it's URL encoded)
    import urllib.parse
    decoded_number = urllib.parse.unquote(whatsapp_number)
    return render_template('player_registration.html', whatsapp_number=decoded_number)

@app.route('/api/register', methods=['POST'])
def api_register_player():
    """Register a new player via the public registration form"""
    try:
        data = request.get_json()
        
        if not data or not data.get('name'):
            return jsonify({'success': False, 'error': 'Player name is required'}), 400
        
        name = data['name'].strip()
        whatsapp = data.get('whatsapp_number', '').strip() or None
        
        # Check if player with same name already exists
        existing_player = Player.query.filter_by(name=name).first()
        if existing_player:
            return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400
        
        # Create new player
        player = Player(
            name=name,
            whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None
        )
        
        db.session.add(player)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Welcome {name}! You have been registered successfully.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/registration-link', methods=['POST'])
@admin_required
def generate_registration_link():
    """Generate a shareable registration link for a player's WhatsApp number"""
    try:
        data = request.get_json()
        player_id = data.get('player_id')

        if not player_id:
            return jsonify({'success': False, 'error': 'Player ID is required'}), 400

        player = Player.query.get(player_id)
        if not player:
            return jsonify({'success': False, 'error': 'Player not found'}), 400

        if not player.whatsapp_number:
            return jsonify({'success': False, 'error': 'Player does not have a WhatsApp number'}), 400

        # Get base URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')

        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"

        # Sanitize the WhatsApp number (remove spaces, dashes, etc.)
        sanitized_whatsapp = sanitize_phone_number(player.whatsapp_number)

        # Create registration link with the sanitized WhatsApp number
        encoded_whatsapp = urllib.parse.quote(sanitized_whatsapp, safe='')
        registration_url = f"{base_url}/register/{encoded_whatsapp}"

        return jsonify({
            'success': True,
            'registration_url': registration_url,
            'whatsapp_number': player.whatsapp_number,
            'player_name': player.name
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/general-registration-link', methods=['POST'])
@admin_required
def generate_general_registration_link():
    """Generate a general registration link for anyone to join"""
    try:
        # Get base URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
        
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        # Create general registration link
        registration_url = f"{base_url}/register"
        
        return jsonify({
            'success': True, 
            'registration_url': registration_url,
            'link_type': 'general'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/dashboard/<token>')
def player_dashboard(token):
    """Player dashboard accessible via token"""
    # Find the pick token
    pick_token = PickToken.query.filter_by(token=token).first()
    
    if not pick_token:
        return render_template('pick_error.html', error="Invalid dashboard link", player_nav_only=True), 404
    
    player = pick_token.player
    current_round = get_current_active_round()

    return render_template('player_dashboard.html', 
                         player=player, 
                         current_round=current_round,
                         token=token,
                         player_nav_only=True)

@app.route('/api/player/<token>/league-table')
def get_player_league_table(token):
    """API endpoint for league table data"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404
    
    try:
        players = Player.query.all()
        league_data = []
        
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            wins = sum(1 for pick in picks if pick.is_winner == True)
            losses = sum(1 for pick in picks if pick.is_winner == False)
            pending = sum(1 for pick in picks if pick.is_winner is None)
            rounds_survived = wins
            
            league_data.append({
                'name': player.name,
                'status': player.status,
                'rounds_survived': rounds_survived,
                'wins': wins,
                'losses': losses,
                'pending': pending,
                'total_picks': len(picks)
            })
        
        # Sort by status priority and then by rounds survived
        status_priority = {'active': 1, 'winner': 2, 'eliminated': 3}
        league_data.sort(key=lambda x: (status_priority.get(x['status'], 4), -x['rounds_survived'], x['name']))
        
        return jsonify({
            'success': True,
            'league_table': league_data
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/player/<token>/pick-history')
def get_player_pick_history(token):
    """API endpoint for player's pick history (current cycle only)"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404

    try:
        player = pick_token.player

        # Get current cycle from active/pending round, or latest round
        current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.id.desc()).first()
        if not current_round:
            current_round = Round.query.order_by(Round.id.desc()).first()
        current_cycle = current_round.cycle_number or 1 if current_round else 1

        # Only show picks from the current cycle (resets after rollover)
        picks = Pick.query.filter_by(player_id=player.id).join(Round).filter(
            Round.cycle_number == current_cycle
        ).order_by(Round.round_number).all()

        pick_history = []
        for pick in picks:
            round_info = db.session.get(Round, pick.round_id)
            pick_history.append({
                'round_number': round_info.round_number,
                'pl_matchday': round_info.pl_matchday,
                'team_picked': pick.team_picked,
                'is_winner': pick.is_winner,
                'timestamp': pick.timestamp.strftime('%Y-%m-%d %H:%M') if pick.timestamp else None,
                'round_status': round_info.status,
                'cycle_number': round_info.cycle_number
            })

        return jsonify({
            'success': True,
            'pick_history': pick_history,
            'player_name': player.name,
            'current_cycle': current_cycle
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/player/<token>/upcoming-fixtures')
def get_player_upcoming_fixtures(token):
    """API endpoint for upcoming fixtures and available teams"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404
    
    try:
        player = pick_token.player
        current_round = get_current_active_round()

        if not current_round:
            return jsonify({
                'success': True,
                'current_round': None,
                'fixtures': [],
                'used_teams': [],
                'has_picked': False
            })
        
        # Get fixtures for current round
        fixtures = Fixture.query.filter_by(round_id=current_round.id).all()

        # Get player's used teams for THIS CYCLE ONLY
        # This ensures teams become available again after a rollover (new cycle)
        current_cycle = current_round.cycle_number or 1
        previous_picks = Pick.query.filter_by(player_id=player.id).join(Round).filter(
            Round.cycle_number == current_cycle
        ).all()
        used_teams = [pick.team_picked for pick in previous_picks]
        
        # Check if player has already picked for current round
        current_pick = Pick.query.filter_by(player_id=player.id, round_id=current_round.id).first()
        
        fixtures_data = []
        for fixture in fixtures:
            fixtures_data.append({
                'home_team': fixture.home_team,
                'away_team': fixture.away_team,
                'date': fixture.date.strftime('%Y-%m-%d') if fixture.date else None,
                'time': fixture.time.strftime('%H:%M') if fixture.time else None,
                'status': fixture.status,
                'home_used': fixture.home_team in used_teams,
                'away_used': fixture.away_team in used_teams
            })
        
        return jsonify({
            'success': True,
            'current_round': {
                'round_number': current_round.round_number,
                'pl_matchday': current_round.pl_matchday,
                'status': current_round.status
            },
            'fixtures': fixtures_data,
            'used_teams': used_teams,
            'has_picked': current_pick is not None,
            'current_pick': current_pick.team_picked if current_pick else None
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Manual WhatsApp Reminder System
class WhatsAppReminder:
    """Class to handle WhatsApp reminder links for manual sending"""
    
    @staticmethod
    def generate_reminder_data(player, round_obj, reminder_type, pick_token):
        """Generate WhatsApp reminder data for manual sending"""
        
        if not player.whatsapp_number:
            return None

        # Determine anchor (kickoff) and cutoff (1 hour before kickoff) times
        anchor_time = getattr(round_obj, 'first_kickoff_at', None) or getattr(round_obj, 'end_date', None)
        cutoff_time = anchor_time - timedelta(hours=1) if anchor_time else None

        def _format_time_remaining(target):
            """Return a friendly countdown like '90 minutes' or '2 hours 15 minutes'."""
            if not target:
                return None
            delta = target - datetime.utcnow()
            total_minutes = int(delta.total_seconds() // 60)
            if total_minutes <= 0:
                return "moments"
            hours, minutes = divmod(total_minutes, 60)
            parts = []
            if hours:
                parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
            if minutes:
                parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")
            return " ".join(parts) if parts else "minutes"

        time_remaining = _format_time_remaining(cutoff_time)
            
        # Get base URL
        base_url = os.environ.get('BASE_URL', 'https://localhost:5000')
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        pick_url = pick_token.get_pick_url(base_url)
        dashboard_url = f"{base_url}/dashboard/{pick_token.token}"
        
        # Customize message based on reminder type
        if reminder_type == '4_hour':
            urgency = "⏰ Reminder"
        elif reminder_type == '2_hour':
            urgency = "⏰ Reminder"
        else:
            urgency = "📝 Reminder"

        if time_remaining:
            time_msg = f"You have about {time_remaining} before the pick window closes (1 hour before kickoff)."
        else:
            time_msg = "Don't forget to submit your pick before the cutoff."
        
        message = f"""{urgency}

Hi {player.name}! 👋

{time_msg} to submit your pick for Round {round_obj.round_number} (PL Matchday {round_obj.pl_matchday}).

Haven't picked yet? Don't get eliminated! 

🎯 Make your pick: {pick_url}

📊 Check your dashboard: {dashboard_url}

Good luck! 🍀
Last Man Standing"""
        
        # Generate WhatsApp link (prefer WhatsApp Web on desktop per request)
        encoded_message = message.replace('\n', '%0A').replace(' ', '%20')
        # Sanitize and clean the number (remove spaces, dashes, then remove +)
        sanitized_number = sanitize_phone_number(player.whatsapp_number)
        clean_number = sanitized_number.replace('+', '')
        # Use WhatsApp Web as default so it opens in the browser
        whatsapp_link = f"https://web.whatsapp.com/send?phone={clean_number}&text={encoded_message}"
        
        return {
            'player_name': player.name,
            'player_id': player.id,
            'whatsapp_number': player.whatsapp_number,
            'message': message,
            'whatsapp_link': whatsapp_link,
            'reminder_type': reminder_type,
            'round_number': round_obj.round_number
        }
    
def get_due_reminders():
    """Get reminders that are due and ready for manual sending"""
    try:
        with app.app_context():
            # Lazy auto-schedule: ensure reminders exist for the active round
            try:
                active_round = get_current_active_round()
                if active_round:
                    ReminderSchedule.create_reminders_for_round(active_round.id)
            except Exception as _e:
                print(f"Auto-schedule skipped/failed: {_e}")

            pending_reminders = ReminderSchedule.get_pending_reminders()
            reminder_data = []
            
            for reminder in pending_reminders:
                # Check if player has already made a pick for this round
                existing_pick = Pick.query.filter_by(
                    player_id=reminder.player_id,
                    round_id=reminder.round_id
                ).first()
                
                if existing_pick:
                    print(f"Player {reminder.player.name} already picked for R{reminder.round.round_number}, marking reminder as sent")
                    reminder.mark_as_sent()
                    continue
                
                # Get or create pick token
                pick_token = PickToken.create_for_player_round(reminder.player_id, reminder.round_id)
                db.session.commit()
                
                # Generate reminder data
                data = WhatsAppReminder.generate_reminder_data(
                    reminder.player,
                    reminder.round,
                    reminder.reminder_type,
                    pick_token
                )
                
                if data:
                    data['reminder_id'] = reminder.id
                    # Provide local-time ISO for accurate browser rendering
                    try:
                        data['scheduled_time'] = to_local(reminder.scheduled_time).isoformat()
                    except Exception:
                        data['scheduled_time'] = reminder.scheduled_time.isoformat()
                    reminder_data.append(data)
                    
            return reminder_data
            
    except Exception as e:
        print(f"Error getting due reminders: {e}")
        return []

# API Routes for reminder management
@app.route('/api/admin/schedule-reminders/<int:round_id>', methods=['POST'])
@admin_required
def schedule_reminders_for_round(round_id):
    """Admin endpoint to manually schedule reminders for a round"""
    try:
        reminders_created = ReminderSchedule.create_reminders_for_round(round_id)
        return jsonify({
            'success': True,
            'message': f'Created {reminders_created} reminders',
            'reminders_created': reminders_created
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/due-reminders')
@admin_required
def get_due_reminders_api():
    """Get all due reminders ready for manual sending"""
    try:
        reminder_data = get_due_reminders()
        return jsonify({
            'success': True,
            'due_reminders': reminder_data,
            'count': len(reminder_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/mark-reminder-sent/<int:reminder_id>', methods=['POST'])
@admin_required
def mark_reminder_sent(reminder_id):
    """Mark a reminder as sent after manual WhatsApp sending"""
    try:
        reminder = ReminderSchedule.query.get(reminder_id)
        if not reminder:
            return jsonify({'success': False, 'error': 'Reminder not found'}), 404
        
        reminder.mark_as_sent()
        return jsonify({
            'success': True,
            'message': 'Reminder marked as sent'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/reminders-dashboard')
@admin_required
def reminders_dashboard():
    """Admin page for managing reminders"""
    current_round = get_current_active_round()
    # Derive first kickoff for display if not stored on the round
    first_kickoff = None
    cutoff_time = None
    try:
        if current_round:
            anchor = current_round.first_kickoff_at or _earliest_kickoff_for_round(current_round) or current_round.end_date
            if anchor:
                first_kickoff = to_local(anchor)
                cutoff_time = to_local(anchor - timedelta(hours=1))
    except Exception:
        pass
    return render_template('reminders_dashboard.html', current_round=current_round, first_kickoff=first_kickoff, cutoff_time=cutoff_time)

@app.route('/admin/statistics')
@admin_required
def admin_statistics_page():
    """Standalone Player Statistics Dashboard page (no JS fetch required)."""
    try:
        # Competition overview
        total_players = Player.query.count()
        active_players = Player.query.filter_by(status='active').count()
        eliminated_players = Player.query.filter_by(status='eliminated').count()
        total_rounds = Round.query.count()
        completed_rounds = Round.query.filter_by(status='completed').count()
        active_round = Round.query.filter_by(status='active').first()

        # Player stats
        players = Player.query.all()
        player_stats = []
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            total_picks = len(picks)
            winning_picks = len([p for p in picks if p.is_winner])
            teams_used = list(set([p.team_picked for p in picks]))
            # Current survival streak
            streak = 0
            for p in reversed(picks):
                if p.is_winner is True:
                    streak += 1
                elif p.is_winner is False:
                    break
            player_stats.append({
                'name': player.name,
                'status': player.status,
                'total_picks': total_picks,
                'winning_picks': winning_picks,
                'success_rate': round((winning_picks / total_picks * 100) if total_picks else 0, 1),
                'current_streak': streak,
            })

        # Pick history
        all_picks = Pick.query.join(Player).join(Round).all()
        pick_history = []
        for pick in all_picks:
            pick_history.append({
                'player_name': pick.player.name,
                'round_number': pick.round.round_number,
                'team_picked': team_abbrev(pick.team_picked),
                'result': 'Winner' if pick.is_winner is True else ('Eliminated' if pick.is_winner is False else 'Pending'),
                'pick_date': pick.timestamp.strftime('%Y-%m-%d %H:%M') if getattr(pick, 'timestamp', None) else 'Unknown'
            })

        competition_stats = {
            'total_players': total_players,
            'active_players': active_players,
            'eliminated_players': eliminated_players,
            'elimination_rate': round((eliminated_players / total_players * 100) if total_players > 0 else 0, 1),
            'total_rounds': total_rounds,
            'completed_rounds': completed_rounds,
            'current_round': active_round.round_number if active_round else None
        }

        # Order player stats: active first, then success rate desc, then name
        def ps_key(p):
            pri = 0 if p['status'] == 'active' else (1 if p['status'] == 'winner' else 2)
            return (pri, -p['success_rate'], p['name'])
        player_stats = sorted(player_stats, key=ps_key)

        # Sort pick history by round then name
        pick_history = sorted(pick_history, key=lambda h: (h['round_number'], h['player_name']))

        return render_template(
            'admin_statistics.html',
            competition_stats=competition_stats,
            player_stats=player_stats,
            pick_history=pick_history
        )
    except Exception as e:
        return render_template('admin_statistics.html', error=str(e), competition_stats={}, player_stats=[], pick_history=[]), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
# --- Public pages ---
@app.route('/rules')
def rules():
    return render_template('rules.html')
